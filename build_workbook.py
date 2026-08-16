"""Builds the Gulf Coast Storm Season Workbook.

    python build_workbook.py

Writes `product/gulf-coast-storm-workbook.xlsx`, ten tabs, no macros. It is the
companion to the storm season binder, and the split between them is the whole
design: **the binder is what you carry, the workbook is what you calculate.**

**Do not rebuild the binder in here.** The countdown, the shutdown sequence, the
re-entry walk, the go bag, what to photograph and the calm-week list are
narrative and sequential, they work better on paper, and duplicating them makes
the bundle look like one product sold twice. What belongs here is arithmetic,
aggregation, running totals against a coverage limit, and a flag that fires when
a number crosses a line. Everything on these ten tabs is one of those.

**The .xlsx is the master and Sheets is the import**, same as the reserve
planner, and for the same reason: a regional or a v2 edition is then an edit to
`workbook_data.py` rather than an afternoon of retyping into a browser.

Three things are worth reading before changing anything.

**A blank policy limit is a third state, not a zero.** This is the trap the
whole product turns on. A buyer who has not entered their jewelry sub-limit must
never be told they are within it, and `='Coverage & Deductibles'!$B$19` on an
empty cell returns 0 rather than blank, which is exactly how that wrong answer
gets produced. Every cross-tab read of a policy figure is wrapped in
`IF(source="","",source)` so the emptiness survives the hop, and the flag
formula then has three branches instead of two. Same reasoning for the wind
deductible: with no deductible type chosen, `D6` returns blank rather than
guessing the flat-amount branch and reporting $0.

**Nothing policy-specific ships with a default.** Coverage limits, deductibles
and sub-limits come off the buyer's own declarations page or they stay empty.
A plausible-looking 2,500 in a sub-limit cell would produce a confident "you're
covered" on the one calculation a buyer would most want to be right. The range
goes in the helper text beside the cell, never in the cell.

**Every formula has to evaluate in both Excel and Google Sheets.** `formula()`
refuses anything outside `ALLOWED`, so the constraint is enforced rather than
remembered. The banned list is the reserve planner's, arrived at the same way.
"""

import os
import re
import sys

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.properties import PageSetupProperties

from binder_pages import (
    CLAIM_NOTE, DAMAGE_NOTE, DISCLAIMER, INVENTORY_WATCH, MITIGATION,
    SECURITY_NOTE, SUPPLY_INTRO)
from workbook_data import (
    ALE_CATEGORIES, CAUSES, CHECK_ANSWERS, CHECKS, FLOOD_SIDE,
    MITIGATION_CATEGORIES, RECEIPT_CATEGORIES, RED_FLAGS, REIMBURSED, REPAIRED,
    ROOMS, SPECIAL, SPECIAL_WITH_NONE, SUBLIMIT_HELP, UNCATEGORIZED,
    WIND_SIDE, WIND_TYPES, YES_NO, check_against_binder, supply_rows)

OUT = os.path.join("product", "gulf-coast-storm-workbook.xlsx")

COV = "'Coverage & Deductibles'"
INV = "'Home Inventory'"
SUP = "'Supply Calculator'"
DMG = "'Damage Log'"
RCP = "'Receipts & Loss of Use'"
CALL = "'Claim Call Log'"
CON = "'Contractor Comparison'"

# Coverage & Deductibles is a fixed block: B4:B16 inputs, D6/D9/D11/D17/D18
# calculations, B19:B24 sub-limits. Everything downstream reads those addresses
# by name, so they are constants rather than literals scattered through the file.
COV_WIND = "$D$6"
COV_CONTENTS = "$D$9"
COV_ALE = "$D$11"
COV_BOTH = "$D$17"
COV_GAP = "$D$18"
COV_FLOOD_DED = "$B$15"
SUBLIMIT_FIRST = 19
SUBLIMIT_LAST = SUBLIMIT_FIRST + len(SPECIAL) - 1

# What every tab says when it needs a policy number the buyer has not entered.
# One sentence, one place to change it, and it names the tab to go to.
UNSET = "Not yet. Enter it on Coverage & Deductibles."

INV_HEAD, INV_FIRST, INV_LAST = 4, 5, 504
SUP_HEAD, SUP_FIRST = 14, 15
SUP_LAST = SUP_FIRST + len(supply_rows()) - 1
DMG_HEAD, DMG_FIRST, DMG_LAST = 10, 11, 210
RCP_HEAD, RCP_FIRST, RCP_LAST = 10, 11, 210
CALL_HEAD, CALL_FIRST, CALL_LAST = 4, 5, 204

# Contractor Comparison, four contractors across C:F.
CHECK_FIRST, CHECK_LAST = 5, 14
FLAG_FIRST, FLAG_LAST = 17, 24
ROW_SCORE, ROW_FLAGS, ROW_VERDICT, ROW_COMPANY = 15, 25, 26, 4
CONTRACTOR_COLS = ["C", "D", "E", "F"]

# Dashboard blocks.
DASH_SUBLIMIT_FIRST = 19
DASH_SUBLIMIT_LAST = DASH_SUBLIMIT_FIRST + len(SPECIAL) - 1
DASH_ROOM_FIRST = 19
DASH_ROOM_LAST = DASH_ROOM_FIRST + len(ROOMS) - 1
DASH_CONTRACTOR_FIRST = 31


def configure(inventory=500, logs=200, calls=200):
    """Build smaller log blocks. Only `qa_workbook.py` uses this.

    The formulas, the references and every summary cell are the ones that ship.
    A pure-Python formula engine will not chew through five hundred inventory
    rows in a sitting, and shrinking the row count rather than the model is what
    keeps the QA run a test of the product instead of a test of a sketch of it.
    """
    global INV_LAST, DMG_LAST, RCP_LAST, CALL_LAST
    INV_LAST = INV_FIRST + inventory - 1
    DMG_LAST = DMG_FIRST + logs - 1
    RCP_LAST = RCP_FIRST + logs - 1
    CALL_LAST = CALL_FIRST + calls - 1


# Where the Dashboard's numbers land. `sheet_dashboard()` checks the label beside
# every one of them at build time, so moving a line without updating this map
# fails the build rather than quietly pointing the QA harness at the wrong cell.
DASH = {
    "wind": ("B5", "Wind deductible, in dollars"),
    "flood": ("B6", "Flood deductible"),
    "both": ("B7", "If wind and flood both apply, you pay both"),
    "gap": ("B8", "What you would have to find"),
    "listed": ("B9", "Contents listed so far"),
    "contents_limit": ("B10", "Your contents limit"),
    "pct": ("B11", "Listed, as a share of that limit"),
    "items": ("B12", "Items logged"),
    "photographed": ("B13", "Photographed"),
    "supply": ("B14", "Supply readiness"),
    "unentered": ("B15", "Sub-limits not yet entered, of 6"),
    "wind_damage": ("B30", "Wind side damage"),
    "flood_damage": ("B31", "Flood side damage"),
    "uncat": ("B32", "Damage with no cause named"),
    "receipts": ("B33", "Total spent since the storm"),
    "ale_pct": ("B34", "Loss of use, share of limit used"),
    "mitigation": ("B35", "Mitigation spending, a separate pot"),
    "days": ("B36", "Days since you last logged a call"),
}

# Every function the workbook is allowed to use. The list is the compatibility
# promise: anything not on it either breaks on the .xlsx round trip or behaves
# differently in one of the two apps.
ALLOWED = {
    "IF", "AND", "OR", "INDEX", "MATCH", "SUMIF", "SUMIFS", "COUNTIF",
    "COUNTIFS", "COUNTA", "IFERROR", "MIN", "MAX", "ROUND", "ROUNDUP",
    "TODAY", "SUM", "SUMPRODUCT", "TEXT",
}

INK = "1A1A1A"           # formula results, the default
INPUT_BLUE = "1F4E79"    # safe to edit
PULLED_GREEN = "1E6B45"  # comes from another tab
MUTED = "6B6B6B"
YELLOW = "FFF2CC"        # a key assumption, or an answer still missing
GRAY = "F2F2F2"          # locked, or unknown
HEAD = "D9E2EC"
BAR = "0F2E4A"
RED_FILL = "F8D0D0"
AMBER_FILL = "FCE4B6"
GREEN_FILL = "D7ECDD"
RULE = Side(style="thin", color="B0BEC5")

MONEY = '"$"#,##0'
PCT = "0%"
DATE = "yyyy-mm-dd"


def formula(text):
    """Return a formula after checking it can survive both apps.

    Catches the two mistakes that are expensive here: an unbalanced paren, which
    makes Excel offer to repair the file and stalls the QA run before it starts,
    and a function outside the compatible set, which produces #NAME? on one
    platform and works fine on the other.
    """
    if text.count("(") != text.count(")"):
        raise SystemExit("unbalanced parens:\n  {0}".format(text))
    for name in re.findall(r"([A-Z][A-Z0-9\.]*)\s*\(", text):
        if name not in ALLOWED:
            raise SystemExit("{0}() is not on the compatible list:\n  {1}".format(name, text))
    return text


def style(cell, *, color=INK, bold=False, fill=None, fmt=None, align=None,
          wrap=False, locked=True, size=None, italic=False):
    cell.font = Font(name="Arial", size=size or 10, bold=bold, italic=italic, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.protection = Protection(locked=locked)
    return cell


def put(ws, ref, value, **kw):
    cell = ws[ref]
    cell.value = value
    return style(cell, **kw)


def heading(ws, ref, text, size=14):
    put(ws, ref, text, bold=True, size=size, color=BAR)


def note(ws, ref, text, span=None, height=None):
    put(ws, ref, text, color=MUTED, italic=True, wrap=True)
    if span:
        ws.merge_cells("{0}:{1}".format(ref, span))
    if height:
        ws.row_dimensions[int(re.sub(r"[A-Z]", "", ref))].height = height


def band(ws, row, headers, height=30):
    """Write a header band and return the row it landed on."""
    for i, text in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=text)
        style(cell, bold=True, fill=HEAD, wrap=True, align="left")
        cell.border = Border(bottom=RULE)
    ws.row_dimensions[row].height = height
    return row


def section(ws, row, text, span):
    """A full-width bar that separates before-the-storm from after."""
    cell = put(ws, "A{0}".format(row), text, bold=True, size=11, color="FFFFFF", fill=BAR)
    ws.merge_cells("A{0}:{1}{0}".format(row, span))
    ws.row_dimensions[row].height = 20
    return cell


def widths(ws, spec):
    for letter, w in spec.items():
        ws.column_dimensions[letter].width = w


def editable(ws, columns, first, last, fmt=None):
    """Open a block of log columns to the buyer.

    The style rides on the column rather than on each cell, which is what keeps
    a blank cell genuinely blank. Writing an empty string into two thousand
    cells to color them would make COUNTA count them.
    """
    for col in columns:
        dim = ws.column_dimensions[col]
        dim.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        dim.protection = Protection(locked=False)
        if fmt and col in fmt:
            dim.number_format = fmt[col]
        dim.customFormat = True
    return first, last


def dropdown(ws, listname, cells):
    dv = DataValidation(type="list", formula1="={0}".format(listname), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cells)
    return dv


def lock(ws):
    """Warning-only protection: on, but with no password.

    A buyer who wants to restructure the sheet clicks Unprotect once and is never
    asked for anything. A locked-down workbook generates support email, and the
    protection here is only there to stop a stray paste over a formula column.
    """
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = False
    ws.protection.autoFilter = False


def printing(ws, area, landscape=False, titles=None):
    """Letter, fit to one page wide, so it three-hole-punches into the binder.

    A buyer who owns both products wants the filled inventory on paper inside
    the binder, because after landfall there is no laptop and no power. Every
    tab prints or the pair does not work.
    """
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    ws.print_area = area
    if titles:
        ws.print_title_rows = titles


def unset(source_ref):
    """Read a policy figure from Coverage & Deductibles without losing its blank.

    `='Coverage & Deductibles'!$B$19` on an empty cell is 0, not blank, and a
    zero limit compares as "you are over it" or "you are within it" depending on
    which way the test runs. Neither is true. This keeps the emptiness.
    """
    return "IF({0}=\"\",\"\",{0})".format(source_ref)


# ------------------------------------------------------------------------ Lists

LIST_COLUMNS = [
    ("Rooms", "RoomList", lambda: ROOMS),
    ("Special categories", "CategoryList", lambda: SPECIAL_WITH_NONE),
    ("Yes / No", "YesNoList", lambda: YES_NO),
    ("Wind deductible type", "WindTypeList", lambda: WIND_TYPES),
    ("Damage cause", "CauseList", lambda: CAUSES),
    ("Repaired", "RepairedList", lambda: REPAIRED),
    ("Receipt category", "ReceiptList", lambda: RECEIPT_CATEGORIES),
    ("Reimbursed", "ReimbursedList", lambda: REIMBURSED),
    ("Yes / No / Pending", "CheckList", lambda: CHECK_ANSWERS),
]


def sheet_lists(wb):
    ws = wb.create_sheet("Lists")
    for i, (title, name, values) in enumerate(LIST_COLUMNS):
        column = i + 1
        style(ws.cell(row=1, column=column, value=title), bold=True)
        items = values()
        for j, text in enumerate(items):
            style(ws.cell(row=2 + j, column=column, value=text))
        # Named ranges rather than raw cross-sheet references. A direct
        # `Lists!$A$2:$A$10` inside a data validation is legal in Excel and lost
        # by some importers; a name survives both.
        letter = ws.cell(row=1, column=column).column_letter
        ref = "Lists!${0}$2:${0}${1}".format(letter, 1 + len(items))
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
        ws.column_dimensions[letter].width = 28
    ws.sheet_state = "hidden"
    lock(ws)
    return ws


# -------------------------------------------------------- Coverage & Deductibles

# row, label, default, list, kind, format, the helper text beside it
COVERAGE_INPUTS = [
    (4, "Dwelling coverage limit ($)", None, None, "input", MONEY,
     "Coverage A on your declarations page. It is what the structure is insured "
     "for, not what the house would sell for."),
    (5, "Standard deductible ($)", None, None, "input", MONEY,
     "The everyday one. It applies to a kitchen fire and a burst pipe, and it is "
     "not the one that applies to a hurricane."),
    (6, "Wind or hurricane deductible type", None, "WindTypeList", "assume", "@",
     "Almost every Gulf policy uses a percentage. Look for the word hurricane, "
     "named storm or windstorm on the declarations page."),
    (7, "Wind deductible, as a percentage", None, None, "assume", "0.0%",
     "Enter 2% as 2%, or 0.02. Commonly 1 to 5 percent of the dwelling limit."),
    (8, "Wind deductible, flat amount ($)", None, None, "assume", MONEY,
     "Only used if you picked Flat dollar amount above. Leave it empty otherwise."),
    (9, "Contents limit ($), if you know it", None, None, "input", MONEY,
     "Coverage C. Leave it empty and the percentage on the next line is used "
     "instead."),
    (10, "...or contents as a share of dwelling", 0.5, None, "assume", PCT,
     "Usually 50 to 70 percent. This is the fallback, and it is a guess about "
     "your policy until you replace it with the real figure above."),
    (11, "Loss of use limit ($), if you know it", None, None, "input", MONEY,
     "Coverage D. It pays for living somewhere else, and it is a separate limit "
     "from the building and the contents."),
    (12, "...or loss of use as a share of dwelling", 0.2, None, "assume", PCT,
     "Usually 20 to 30 percent. Same warning as the line above."),
    (13, "Flood building limit ($)", None, None, "input", MONEY,
     "Off the flood policy, which is a different document from the homeowners "
     "one. NFIP caps the building at $250,000."),
    (14, "Flood contents limit ($)", None, None, "input", MONEY,
     "Flood contents is optional and often not bought. NFIP caps it at $100,000. "
     "If you do not have it, enter 0 and know that."),
    (15, "Flood deductible ($)", None, None, "input", MONEY,
     "A separate deductible on a separate policy, paid on top of the wind one "
     "when both apply."),
    (16, "Your accessible emergency savings ($)", None, None, "input", MONEY,
     "What you could actually reach in a week. Not a retirement account, and not "
     "a credit limit."),
]


def sheet_coverage(wb):
    ws = wb.create_sheet("Coverage & Deductibles")
    heading(ws, "A1", "Coverage & Deductibles")
    note(ws, "A2",
         "Fifteen minutes with your declarations page, and this tab answers the one "
         "question most homeowners have never looked up: what a hurricane costs you "
         "before the policy pays anything. Nothing here is pre-filled, because a "
         "plausible guess about your policy is worse than an empty cell.",
         span="F2", height=42)

    band(ws, 3, ["Off your declarations page", "Your number", "Where to find it",
                 "What it works out to", "What that means"])

    for row, label, default, listname, kind, fmt, why in COVERAGE_INPUTS:
        put(ws, "A{0}".format(row), label, bold=True)
        cell = put(ws, "B{0}".format(row), default, color=INPUT_BLUE, fmt=fmt,
                   fill=YELLOW if kind == "assume" else None, locked=False)
        cell.border = Border(bottom=RULE)
        note(ws, "C{0}".format(row), why)
        ws.row_dimensions[row].height = 30
        if listname:
            dropdown(ws, listname, cell)

    # D6. With no deductible type chosen this refuses to answer rather than
    # falling through to the flat-amount branch and reporting a $0 wind
    # deductible, which is the most dangerous wrong number the workbook could
    # print. Everything downstream tests for the blank.
    put(ws, "D6", formula(
        '=IF($B$6="","",IF($B$6="{0}",$B$4*$B$7,$B$8))'.format(WIND_TYPES[0])),
        bold=True, size=16, fmt=MONEY)
    put(ws, "E6", "Cash out of your pocket before the wind policy pays anything. "
                  "On a $400,000 house a 2 percent deductible is $8,000.",
        wrap=True)

    put(ws, "D9", formula('=IF($B$9<>"",$B$9,$B$4*$B$10)'), fmt=MONEY, bold=True)
    put(ws, "E9", "Contents limit in use. The Home Inventory tab measures your "
                  "listed value against this.", wrap=True)

    put(ws, "D11", formula('=IF($B$11<>"",$B$11,$B$4*$B$12)'), fmt=MONEY, bold=True)
    put(ws, "E11", "Loss of use limit in use. Receipts & Loss of Use measures "
                   "against this.", wrap=True)

    put(ws, "A17", "Wind and flood together", bold=True)
    put(ws, "D17", formula('=IF($D$6="","",$D$6+$B$15)'), bold=True, size=16, fmt=MONEY)
    put(ws, "E17", "Wind and flood are separate policies with separate deductibles. "
                   "A storm that breaks a window and pushes water across the yard "
                   "triggers both, and you pay both.", wrap=True)
    ws.row_dimensions[17].height = 34

    put(ws, "A18", "What you would have to find", bold=True)
    put(ws, "D18", formula('=IF($D$17="","",MAX(0,$D$17-$B$16))'),
        bold=True, size=16, fmt=MONEY)
    put(ws, "E18", "Your combined deductibles, less the savings you could reach. "
                   "If this is not zero, it is the number to work on between now "
                   "and the next season.", wrap=True)
    ws.row_dimensions[18].height = 34

    ws.conditional_formatting.add("D18", CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))

    # The sub-limits. Blank on purpose, yellow so they read as unfinished, and
    # validated to refuse a zero, because "0" and "I have not looked" are
    # different answers and only one of them is safe to act on.
    for i, name in enumerate(SPECIAL):
        row = SUBLIMIT_FIRST + i
        put(ws, "A{0}".format(row), "{0}, sub-limit ($)".format(name), bold=True)
        cell = put(ws, "B{0}".format(row), None, color=INPUT_BLUE, fmt=MONEY,
                   fill=YELLOW, locked=False)
        cell.border = Border(bottom=RULE)
        note(ws, "C{0}".format(row), SUBLIMIT_HELP)
        ws.row_dimensions[row].height = 28
    dv = DataValidation(type="decimal", operator="greaterThan", formula1="0",
                        allow_blank=True, showErrorMessage=True,
                        errorTitle="That has to be a dollar figure",
                        error="Enter the sub-limit from your declarations page. "
                              "Leave it empty if you have not looked it up yet.")
    ws.add_data_validation(dv)
    dv.add("B{0}:B{1}".format(SUBLIMIT_FIRST, SUBLIMIT_LAST))

    note(ws, "E{0}".format(SUBLIMIT_FIRST),
         INVENTORY_WATCH + " Leave these empty until you have read your own policy. "
         "The Dashboard will keep telling you they are missing, which is the "
         "correct behavior: an unknown limit is not the same as a limit you are "
         "inside of.",
         span="F{0}".format(SUBLIMIT_LAST))

    widths(ws, {"A": 38, "B": 16, "C": 46, "D": 18, "E": 46, "F": 4})
    printing(ws, "A1:E{0}".format(SUBLIMIT_LAST))
    lock(ws)
    return ws


# ------------------------------------------------------------------ Home Inventory

INVENTORY_HEADERS = [
    "Room", "Item", "Make & model", "Serial or ID", "Year", "Qty",
    "Replacement cost today, each ($)", "Line total ($)", "Special category",
    "Photographed?", "Photo reference", "Notes",
]


def sheet_inventory(wb):
    ws = wb.create_sheet("Home Inventory")

    total = "SUM($H${0}:$H${1})".format(INV_FIRST, INV_LAST)
    logged = "COUNTA($B${0}:$B${1})".format(INV_FIRST, INV_LAST)
    share = "IFERROR({0}/{1}!{2},0)".format(total, COV, COV_CONTENTS)

    ws.merge_cells("A1:B2")
    heading(ws, "A1", "Home Inventory")

    summary = [
        ("C", "Total listed value", "=" + total, MONEY),
        ("E", "Your contents limit", formula("={0}!{1}".format(COV, COV_CONTENTS)), MONEY),
        ("G", "Listed, as a share of it", formula("=" + share), PCT),
        ("I", "Items logged", formula("=" + logged), "0"),
        ("K", "Photographed", formula(
            '=IFERROR(COUNTIF($J${0}:$J${1},"Yes")/{2},0)'.format(
                INV_FIRST, INV_LAST, logged)), PCT),
    ]
    for col, label, value, fmt in summary:
        nxt = chr(ord(col) + 1)
        put(ws, "{0}1".format(col), label, bold=True, color=MUTED, align="center")
        ws.merge_cells("{0}1:{1}1".format(col, nxt))
        put(ws, "{0}2".format(col), formula(value) if value.startswith("=") else value,
            bold=True, size=14, fmt=fmt, align="center")
        ws.merge_cells("{0}2:{1}2".format(col, nxt))
    ws.row_dimensions[2].height = 22

    # The under-listing nudge. Four states, and the middle one is the point:
    # most households under-list by a wide margin, and a product that
    # congratulated them at 40 percent of their limit would be doing them harm.
    put(ws, "A3", formula(
        '=IF({logged}=0,'
        '"Start with the expensive things. Twenty items at $500 and up matter more '
        'than two hundred at $20.",'
        'IF({share}<0.5,'
        '"You have listed less than half your contents limit. Most households '
        'under-list by a wide margin. Keep going, one room at a time.",'
        'IF({share}>1,'
        '"You have listed more than your contents limit. Worth a conversation with '
        'your agent before renewal.",'
        '"Good coverage of your limit. Re-walk the house each spring.")))'.format(
            logged=logged, share=share)),
        wrap=True, fill=YELLOW, color=BAR, bold=True)
    ws.merge_cells("A3:L3")
    ws.row_dimensions[3].height = 30

    band(ws, INV_HEAD, INVENTORY_HEADERS)

    editable(ws, list("ABCDEFG") + list("IJKL"), INV_FIRST, INV_LAST,
             fmt={"E": "0", "F": "0", "G": MONEY})
    for row in range(INV_FIRST, INV_LAST + 1):
        put(ws, "H{0}".format(row), formula(
            '=IF($B{0}="","",$F{0}*$G{0})'.format(row)), fmt=MONEY)

    dropdown(ws, "RoomList", "A{0}:A{1}".format(INV_FIRST, INV_LAST))
    dropdown(ws, "CategoryList", "I{0}:I{1}".format(INV_FIRST, INV_LAST))
    dropdown(ws, "YesNoList", "J{0}:J{1}".format(INV_FIRST, INV_LAST))

    widths(ws, {"A": 30, "B": 32, "C": 24, "D": 20, "E": 8, "F": 6, "G": 16,
                "H": 14, "I": 20, "J": 14, "K": 18, "L": 30})
    ws.freeze_panes = "A{0}".format(INV_FIRST)
    printing(ws, "A1:L{0}".format(INV_LAST), landscape=True,
             titles="{0}:{0}".format(INV_HEAD))
    lock(ws)
    return ws


# --------------------------------------------------------------- Supply Calculator

SUPPLY_INPUTS = [
    (4, "People in the house", None, None, "input", "0",
     "Everyone who will be riding it out here, including anyone staying with you."),
    (5, "Pets, in large-dog equivalents", None, None, "input", "0",
     "A large dog is 1. Two cats are about 1 between them. This only drives water "
     "and food."),
    (6, "Days of supply", 7, None, "assume", "0",
     "Seven, and the binder says why: federal guidance is three days, and three "
     "days is for a house that was merely inconvenienced. Fourteen if you are "
     "rural, on a well, or at the end of a long feeder line."),
    (7, "Coolers you actually own", None, None, "input", "0",
     "Ice is one bag per cooler per day, and it will not be for sale afterward."),
    (8, "Vehicles to fuel", None, None, "input", "0",
     "Every tank full at 72 hours. Stations run dry and the pumps are electric."),
    (9, "Generator?", None, "YesNoList", "input", "@",
     "Says whether the fuel line below calculates at all."),
    (10, "Generator hours per day", 9, None, "assume", "0",
     "A portable running eight to ten hours a day burns roughly 5 gallons. Change "
     "this and the fuel figure scales with it."),
]


def sheet_supply(wb):
    ws = wb.create_sheet("Supply Calculator")
    heading(ws, "A1", "Supply Calculator")
    note(ws, "A2", SUPPLY_INTRO, span="E3", height=44)

    for row, label, default, listname, kind, fmt, why in SUPPLY_INPUTS:
        put(ws, "A{0}".format(row), label, bold=True)
        cell = put(ws, "B{0}".format(row), default, color=INPUT_BLUE, fmt=fmt,
                   fill=YELLOW if kind == "assume" else None, locked=False)
        cell.border = Border(bottom=RULE)
        note(ws, "C{0}".format(row), why)
        ws.merge_cells("C{0}:E{0}".format(row))
        ws.row_dimensions[row].height = 30
        if listname:
            dropdown(ws, listname, cell)

    put(ws, "A12", "Supply readiness", bold=True)
    put(ws, "B12", formula(
        '=IFERROR(COUNTIFS($C${0}:$C${1},">0",$E${0}:$E${1},"=0")'
        '/COUNTIF($C${0}:$C${1},">0"),0)'.format(SUP_FIRST, SUP_LAST)),
        bold=True, size=14, fmt=PCT)
    note(ws, "C12", "The share of the lines below that you have already covered. "
                    "It is the number the Dashboard carries.")
    ws.merge_cells("C12:E12")

    band(ws, SUP_HEAD, ["What", "The rule", "Your number", "Already have", "Still need"])

    for i, (label, rule, unit, value, fmt) in enumerate(supply_rows()):
        row = SUP_FIRST + i
        put(ws, "A{0}".format(row), label, bold=True)
        note(ws, "B{0}".format(row), "{0}. Measured in {1}.".format(rule, unit))
        put(ws, "C{0}".format(row),
            formula(value) if isinstance(value, str) else value, fmt=fmt)
        put(ws, "D{0}".format(row), None, color=INPUT_BLUE, fmt=fmt, locked=False)
        put(ws, "E{0}".format(row), formula(
            "=MAX(0,$C{0}-$D{0})".format(row)), fmt=fmt, bold=True)
        ws.row_dimensions[row].height = 32

    ws.conditional_formatting.add(
        "E{0}:E{1}".format(SUP_FIRST, SUP_LAST),
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", bgColor=AMBER_FILL)))

    note(ws, "A{0}".format(SUP_LAST + 2),
         "Still need is your shopping list. Print this tab and take it, because the "
         "store you are going to will have no power for its own signage either. "
         "Prescriptions and batteries are deliberately not on here: they are "
         "\"check every bottle\" and \"count them\", not arithmetic, and they are "
         "on the binder's page where they belong.",
         span="E{0}".format(SUP_LAST + 3))

    widths(ws, {"A": 30, "B": 46, "C": 14, "D": 14, "E": 14})
    printing(ws, "A1:E{0}".format(SUP_LAST + 3))
    lock(ws)
    return ws


# --------------------------------------------------------------------- Damage Log

DAMAGE_HEADERS = ["Date", "Room or area", "What is damaged", "Cause",
                  "Photo reference", "Estimated value ($)", "Repaired?", "Notes"]


def _side_total(names):
    """Sum the value column for a set of causes.

    One exact-match SUMIF per cause, added together, rather than one SUMIF with
    a wildcard. "Wind" and "Wind-driven rain" both begin with Wind, so a
    `"Wind*"` criterion would fold the second into the first and quietly
    overstate the side of the claim the buyer is about to argue about.
    """
    parts = ['SUMIF($D${0}:$D${1},"{2}",$F${0}:$F${1})'.format(DMG_FIRST, DMG_LAST, name)
             for name in names]
    return "(" + "+".join(parts) + ")"


def sheet_damage(wb):
    ws = wb.create_sheet("Damage Log")
    heading(ws, "A1", "Damage Log")
    note(ws, "A2", DAMAGE_NOTE, span="H3", height=44)

    band(ws, 5, ["Which policy pays", "Total logged", "Against your deductible"])

    wind = _side_total(WIND_SIDE)
    flood = _side_total(FLOOD_SIDE)

    put(ws, "A6", "Wind side (wind, wind-driven rain, tree)", bold=True)
    put(ws, "B6", formula("=" + wind), bold=True, size=14, fmt=MONEY)
    put(ws, "C6", formula(
        '=IF({0}!{1}="","{2}",IF($B$6>{0}!{1},'
        '"Above your wind deductible","Below your wind deductible, currently no payout"))'
        .format(COV, COV_WIND, UNSET)), wrap=True)

    put(ws, "A7", "Flood side (flood, surge)", bold=True)
    put(ws, "B7", formula("=" + flood), bold=True, size=14, fmt=MONEY)
    put(ws, "C7", formula(
        '=IF({0}!{1}="","{2}",IF($B$7>{0}!{1},'
        '"Above your flood deductible","Below your flood deductible, currently no payout"))'
        .format(COV, COV_FLOOD_DED, UNSET)), wrap=True)

    put(ws, "A8", "No cause named yet", bold=True)
    put(ws, "B8", formula('=SUMIF($D${0}:$D${1},"{2}",$F${0}:$F${1})'.format(
        DMG_FIRST, DMG_LAST, UNCATEGORIZED)), bold=True, size=14, fmt=MONEY)
    put(ws, "C8", "Counts toward neither side, on purpose. Write what you saw and "
                  "let the adjuster reach the conclusion.", wrap=True)
    for row in (6, 7, 8):
        ws.row_dimensions[row].height = 30

    for row, which in ((6, "wind"), (7, "flood")):
        ws.conditional_formatting.add("C{0}".format(row), CellIsRule(
            operator="equal", formula=['"Above your {0} deductible"'.format(which)],
            fill=PatternFill("solid", bgColor=AMBER_FILL)))

    band(ws, DMG_HEAD, DAMAGE_HEADERS)
    editable(ws, list("ABCDEFGH"), DMG_FIRST, DMG_LAST, fmt={"A": DATE, "F": MONEY})
    dropdown(ws, "CauseList", "D{0}:D{1}".format(DMG_FIRST, DMG_LAST))
    dropdown(ws, "RepairedList", "G{0}:G{1}".format(DMG_FIRST, DMG_LAST))

    widths(ws, {"A": 13, "B": 24, "C": 40, "D": 24, "E": 18, "F": 16, "G": 14, "H": 36})
    ws.freeze_panes = "A{0}".format(DMG_FIRST)
    printing(ws, "A1:H{0}".format(DMG_LAST), landscape=True,
             titles="{0}:{0}".format(DMG_HEAD))
    lock(ws)
    return ws


# ------------------------------------------------------ Receipts & Loss of Use

RECEIPT_HEADERS = ["Date", "Category", "Vendor", "Amount ($)", "Paid by",
                   "Receipt image reference", "Submitted to insurer?",
                   "Reimbursed?", "Amount reimbursed ($)"]


def _receipt_total(names):
    parts = ['SUMIF($B${0}:$B${1},"{2}",$D${0}:$D${1})'.format(RCP_FIRST, RCP_LAST, name)
             for name in names]
    # Parenthesized, always. These get used as the numerator of the share-of-limit
    # calculation, and `a+b+c/limit` divides only c. That mistake produces a
    # number, not an error, which is the kind this file exists to stop.
    return "(" + "+".join(parts) + ")"


def sheet_receipts(wb):
    ws = wb.create_sheet("Receipts & Loss of Use")
    heading(ws, "A1", "Receipts & Loss of Use")
    note(ws, "A2",
         "This is the page the binder does not have, and it is the one people lose "
         "money on. Loss of use pays for living somewhere else and draws on its own "
         "limit. Mitigation spending, the tarps and the fans and the generator fuel, "
         "is reimbursable under your duty to prevent further damage and is NOT drawn "
         "from that limit. They are two different pots and they are constantly "
         "conflated. " + MITIGATION,
         span="I3", height=58)

    ale = _receipt_total(ALE_CATEGORIES)
    mitigation = _receipt_total(MITIGATION_CATEGORIES)
    spent = "SUM($D${0}:$D${1})".format(RCP_FIRST, RCP_LAST)
    reimbursed = "SUM($I${0}:$I${1})".format(RCP_FIRST, RCP_LAST)

    totals = [
        ("A5", "Loss of use subtotal", "B5", "=" + ale, MONEY),
        ("D5", "Your loss of use limit", "E5", "={0}!{1}".format(COV, COV_ALE), MONEY),
        ("A6", "Loss of use, share of limit used", "B6",
         "=IFERROR({0}/{1}!{2},0)".format(ale, COV, COV_ALE), PCT),
        ("D6", "Mitigation subtotal, a separate pot", "E6", "=" + mitigation, MONEY),
        ("A7", "Total spent", "B7", "=" + spent, MONEY),
        ("D7", "Total submitted to the insurer", "E7",
         '=SUMIF($G${0}:$G${1},"Yes",$D${0}:$D${1})'.format(RCP_FIRST, RCP_LAST), MONEY),
        ("A8", "Total reimbursed", "B8", "=" + reimbursed, MONEY),
        ("D8", "Still outstanding", "E8", "={0}-{1}".format(spent, reimbursed), MONEY),
    ]
    for label_ref, label, value_ref, value, fmt in totals:
        put(ws, label_ref, label, bold=True)
        put(ws, value_ref, formula(value), bold=True, size=12, fmt=fmt)

    ws.conditional_formatting.add("B6", CellIsRule(
        operator="greaterThan", formula=["0.8"],
        fill=PatternFill("solid", bgColor=AMBER_FILL)))

    band(ws, RCP_HEAD, RECEIPT_HEADERS)
    editable(ws, list("ABCDEFGHI"), RCP_FIRST, RCP_LAST,
             fmt={"A": DATE, "D": MONEY, "I": MONEY})
    dropdown(ws, "ReceiptList", "B{0}:B{1}".format(RCP_FIRST, RCP_LAST))
    dropdown(ws, "YesNoList", "G{0}:G{1}".format(RCP_FIRST, RCP_LAST))
    dropdown(ws, "ReimbursedList", "H{0}:H{1}".format(RCP_FIRST, RCP_LAST))

    widths(ws, {"A": 13, "B": 22, "C": 26, "D": 14, "E": 18, "F": 24, "G": 16,
                "H": 14, "I": 18})
    ws.freeze_panes = "A{0}".format(RCP_FIRST)
    printing(ws, "A1:I{0}".format(RCP_LAST), landscape=True,
             titles="{0}:{0}".format(RCP_HEAD))
    lock(ws)
    return ws


# ---------------------------------------------------------------- Claim Call Log

CALL_HEADERS = ["Date", "Time", "Company", "Name", "Title",
                "Direct number or extension", "Claim number",
                "What was said, and what was promised", "Follow-up due", "Done?"]


def sheet_calls(wb):
    ws = wb.create_sheet("Claim Call Log")
    heading(ws, "A1", "Claim Call Log")

    put(ws, "A2", "Days since your last logged contact", bold=True)
    put(ws, "B2", formula(
        '=IF(COUNTA($A${0}:$A${1})=0,"",TODAY()-MAX($A${0}:$A${1}))'.format(
            CALL_FIRST, CALL_LAST)), bold=True, size=14, fmt="0")
    note(ws, "C2", "Amber past a week, red past a fortnight. A quiet claim is not a "
                   "claim being worked on.")
    ws.merge_cells("C2:F2")
    ws.conditional_formatting.add("B2", CellIsRule(
        operator="greaterThan", formula=["14"],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add("B2", CellIsRule(
        operator="greaterThan", formula=["7"],
        fill=PatternFill("solid", bgColor=AMBER_FILL)))

    note(ws, "A3", CLAIM_NOTE, span="J3", height=44)

    band(ws, CALL_HEAD, CALL_HEADERS)
    editable(ws, list("ABCDEFGHIJ"), CALL_FIRST, CALL_LAST,
             fmt={"A": DATE, "I": DATE})
    dropdown(ws, "YesNoList", "J{0}:J{1}".format(CALL_FIRST, CALL_LAST))

    widths(ws, {"A": 13, "B": 9, "C": 22, "D": 20, "E": 18, "F": 20, "G": 18,
                "H": 46, "I": 14, "J": 10})
    ws.freeze_panes = "A{0}".format(CALL_FIRST)
    printing(ws, "A1:J{0}".format(CALL_LAST), landscape=True,
             titles="{0}:{0}".format(CALL_HEAD))
    lock(ws)
    return ws


# -------------------------------------------------------- Contractor Comparison

def sheet_contractors(wb):
    ws = wb.create_sheet("Contractor Comparison")
    heading(ws, "A1", "Contractor Comparison")
    note(ws, "A2",
         "Ten checks, four contractors, one score each. Any single red flag below "
         "overrides a perfect score, because these are not points on a scale: they "
         "are the ways the storm-chasing crews take money out of a neighborhood "
         "after a landfall, and one of them is enough.",
         span="F2", height=44)

    band(ws, 3, ["The check", "Why it matters"] +
         ["Contractor {0}".format(i + 1) for i in range(len(CONTRACTOR_COLS))])

    put(ws, "A{0}".format(ROW_COMPANY), "Company name", bold=True)
    note(ws, "B{0}".format(ROW_COMPANY), "Fill this in first. Everything below is "
                                         "Yes, No or Pending.")
    for col in CONTRACTOR_COLS:
        put(ws, "{0}{1}".format(col, ROW_COMPANY), None, color=INPUT_BLUE, locked=False)

    for i, (label, why) in enumerate(CHECKS):
        row = CHECK_FIRST + i
        put(ws, "A{0}".format(row), label, bold=True, wrap=True)
        note(ws, "B{0}".format(row), why)
        for col in CONTRACTOR_COLS:
            put(ws, "{0}{1}".format(col, row), None, color=INPUT_BLUE,
                align="center", locked=False)
        ws.row_dimensions[row].height = 34
    dropdown(ws, "CheckList", "{0}{1}:{2}{3}".format(
        CONTRACTOR_COLS[0], CHECK_FIRST, CONTRACTOR_COLS[-1], CHECK_LAST))

    put(ws, "A{0}".format(ROW_SCORE), "Score, out of 10", bold=True, size=12)
    for col in CONTRACTOR_COLS:
        put(ws, "{0}{1}".format(col, ROW_SCORE), formula(
            '=COUNTIF({0}{1}:{0}{2},"Yes")'.format(col, CHECK_FIRST, CHECK_LAST)),
            bold=True, size=14, align="center")

    section(ws, ROW_FLAGS - len(RED_FLAGS) - 1,
            "Red flags. Any one of these ends it, whatever the score says.", "F")

    for i, label in enumerate(RED_FLAGS):
        row = FLAG_FIRST + i
        put(ws, "A{0}".format(row), label, bold=True, wrap=True)
        for col in CONTRACTOR_COLS:
            put(ws, "{0}{1}".format(col, row), None, color=INPUT_BLUE,
                align="center", locked=False)
        ws.row_dimensions[row].height = 26
    dropdown(ws, "YesNoList", "{0}{1}:{2}{3}".format(
        CONTRACTOR_COLS[0], FLAG_FIRST, CONTRACTOR_COLS[-1], FLAG_LAST))

    put(ws, "A{0}".format(ROW_FLAGS), "Red flags counted", bold=True, size=12)
    put(ws, "A{0}".format(ROW_VERDICT), "Verdict", bold=True, size=12)
    for col in CONTRACTOR_COLS:
        put(ws, "{0}{1}".format(col, ROW_FLAGS), formula(
            '=COUNTIF({0}{1}:{0}{2},"Yes")'.format(col, FLAG_FIRST, FLAG_LAST)),
            bold=True, size=14, align="center")
        # A red flag outranks the score. That is the correct logic and the
        # formula is written so it behaves that way rather than averaging.
        put(ws, "{0}{1}".format(col, ROW_VERDICT), formula(
            '=IF({0}{1}>0,"STOP, walk away",'
            'IF({0}{2}>=9,"Cleared",'
            'IF({0}{2}>=6,"Incomplete, finish the checks","Not vetted")))'.format(
                col, ROW_FLAGS, ROW_SCORE)), bold=True, size=12, align="center")

    verdicts = "{0}{1}:{2}{1}".format(CONTRACTOR_COLS[0], ROW_VERDICT, CONTRACTOR_COLS[-1])
    ws.conditional_formatting.add(verdicts, CellIsRule(
        operator="equal", formula=['"STOP, walk away"'],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(verdicts, CellIsRule(
        operator="equal", formula=['"Cleared"'],
        fill=PatternFill("solid", bgColor=GREEN_FILL)))

    widths(ws, {"A": 40, "B": 52, "C": 16, "D": 16, "E": 16, "F": 16})
    printing(ws, "A1:F{0}".format(ROW_VERDICT), titles="3:3")
    lock(ws)
    return ws


# ---------------------------------------------------------------------- Dashboard

def sheet_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    heading(ws, "A1", "Dashboard")
    note(ws, "A2",
         "Everything that matters, on one screen. Every number here is calculated "
         "somewhere else; nothing on this tab is typed. If a figure reads as not "
         "entered, that is the tab to go and open.",
         span="E2", height=28)

    section(ws, 4, "Before the storm", "E")

    inv_total = "{0}!$C$2".format(INV)
    rows = [
        ("wind", "={0}!{1}".format(COV, COV_WIND), MONEY,
         "Cash out of pocket before the wind policy pays anything."),
        ("flood", "={0}".format(unset("{0}!{1}".format(COV, COV_FLOOD_DED))), MONEY,
         "A separate policy, and a separate deductible."),
        ("both", "={0}!{1}".format(COV, COV_BOTH), MONEY,
         "One storm can trigger both. Then you pay both."),
        ("gap", "={0}!{1}".format(COV, COV_GAP), MONEY,
         "Your combined deductibles, less the savings you could reach this week."),
        ("listed", "={0}".format(inv_total), MONEY,
         "What you have actually written down on Home Inventory."),
        ("contents_limit", "={0}!{1}".format(COV, COV_CONTENTS), MONEY,
         "What the policy would pay for contents, in total."),
        ("pct", "=IFERROR({0}/{1}!{2},0)".format(inv_total, COV, COV_CONTENTS), PCT,
         "Most households sit far below this and think they are done."),
        ("items", "={0}!$I$2".format(INV), "0",
         "Lines on the inventory with something written in them."),
        ("photographed", "={0}!$K$2".format(INV), PCT,
         "A photograph of the serial plate is where the make, the model and the "
         "number all come from."),
        ("supply", "={0}!$B$12".format(SUP), PCT,
         "The share of the supply lines you have already covered."),
        ("unentered", "={2}-COUNTIF($C${0}:$C${1},\">0\")".format(
            DASH_SUBLIMIT_FIRST, DASH_SUBLIMIT_LAST, len(SPECIAL)), "0",
         "Each one is a category where you could be capped far below your contents "
         "limit and not know it. On a fresh copy this reads 6."),
    ]
    for key, value, fmt, meaning in rows:
        ref, label = DASH[key]
        row = ref[1:]
        put(ws, "A{0}".format(row), label, bold=True)
        put(ws, ref, formula(value), bold=True, size=14, fmt=fmt, color=PULLED_GREEN)
        note(ws, "C{0}".format(row), meaning)
        ws.merge_cells("C{0}:E{0}".format(row))
        ws.row_dimensions[int(row)].height = 24

    ws.conditional_formatting.add(DASH["gap"][0], CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(DASH["pct"][0], DataBarRule(
        start_type="num", start_value=0, end_type="num", end_value=1,
        color=PULLED_GREEN, showValue=True))
    ws.conditional_formatting.add(DASH["unentered"][0], CellIsRule(
        operator="greaterThan", formula=["0"],
        fill=PatternFill("solid", bgColor=AMBER_FILL)))

    # --- sub-limit exposure, the table the whole "blank is not zero" rule is for
    section(ws, 17, "Special category exposure", "E")
    band(ws, 18, ["Category", "Value owned", "Policy limit", "Exposure",
                  "What that means"])
    for i, name in enumerate(SPECIAL):
        row = DASH_SUBLIMIT_FIRST + i
        limit_ref = "{0}!$B${1}".format(COV, SUBLIMIT_FIRST + i)
        put(ws, "A{0}".format(row), name, bold=True)
        put(ws, "B{0}".format(row), formula(
            '=SUMIF({0}!$I${1}:$I${2},$A{3},{0}!$H${1}:$H${2})'.format(
                INV, INV_FIRST, INV_LAST, row)), fmt=MONEY)
        put(ws, "C{0}".format(row), formula("=" + unset(limit_ref)),
            fmt=MONEY, color=PULLED_GREEN)
        put(ws, "D{0}".format(row), formula(
            '=IF($C{0}="","",MAX(0,$B{0}-$C{0}))'.format(row)), fmt=MONEY, bold=True)
        # Three states, not two. A blank limit must never render as "Within
        # limit", and the wording escalates when the buyer owns something in the
        # category, because that is when not knowing is expensive.
        put(ws, "E{0}".format(row), formula(
            '=IF($C{0}="",'
            'IF($B{0}>0,'
            '"LIMIT NOT ENTERED. You own $"&TEXT($B{0},"#,##0")&'
            '" here and have not entered your limit.",'
            '"Limit not entered"),'
            'IF($B{0}-$C{0}>0,'
            '"OVER SUB-LIMIT. Ask your agent about a scheduled endorsement.",'
            '"Within limit"))'.format(row)), wrap=True, fill=GRAY)
        ws.row_dimensions[row].height = 26

    exposure_block = "A{0}:E{1}".format(DASH_SUBLIMIT_FIRST, DASH_SUBLIMIT_LAST)
    ws.conditional_formatting.add(exposure_block, FormulaRule(
        formula=['AND($C{0}<>"",$B{0}>$C{0})'.format(DASH_SUBLIMIT_FIRST)],
        fill=PatternFill("solid", bgColor=RED_FILL), stopIfTrue=True))
    ws.conditional_formatting.add(exposure_block, FormulaRule(
        formula=['AND($C{0}="",$B{0}>0)'.format(DASH_SUBLIMIT_FIRST)],
        fill=PatternFill("solid", bgColor=AMBER_FILL), stopIfTrue=True))
    ws.conditional_formatting.add(exposure_block, FormulaRule(
        formula=['AND($C{0}<>"",$B{0}<=$C{0})'.format(DASH_SUBLIMIT_FIRST)],
        fill=PatternFill("solid", bgColor=GREEN_FILL), stopIfTrue=True))

    # --- value by room, beside it
    put(ws, "G17", "Value listed by room", bold=True, size=11, color="FFFFFF", fill=BAR)
    ws.merge_cells("G17:H17")
    band_cells = ["Room", "Value listed"]
    for i, text in enumerate(band_cells):
        cell = ws.cell(row=18, column=7 + i, value=text)
        style(cell, bold=True, fill=HEAD, wrap=True, align="left")
        cell.border = Border(bottom=RULE)
    for i, name in enumerate(ROOMS):
        row = DASH_ROOM_FIRST + i
        put(ws, "G{0}".format(row), name)
        put(ws, "H{0}".format(row), formula(
            '=SUMIF({0}!$A${1}:$A${2},$G{3},{0}!$H${1}:$H${2})'.format(
                INV, INV_FIRST, INV_LAST, row)), fmt=MONEY)

    # --- after the storm
    section(ws, 29, "After the storm", "E")
    after = [
        ("wind_damage", "={0}!$B$6".format(DMG), MONEY,
         "Wind, wind-driven rain and tree damage, which is the homeowners policy."),
        ("flood_damage", "={0}!$B$7".format(DMG), MONEY,
         "Flood and surge, which is the flood policy and a different company."),
        ("uncat", "={0}!$B$8".format(DMG), MONEY,
         "Logged with no cause named. Counts toward neither side, on purpose."),
        ("receipts", "={0}!$B$7".format(RCP), MONEY,
         "Everything on Receipts & Loss of Use, both pots together."),
        ("ale_pct", "={0}!$B$6".format(RCP), PCT,
         "Loss of use only. Mitigation is not drawn from this limit."),
        ("mitigation", "={0}!$E$6".format(RCP), MONEY,
         "Tarps, fans, fuel. Reimbursable under your duty to prevent further damage."),
        ("days", "={0}!$B$2".format(CALL), "0",
         "Since the last line on the Claim Call Log. Amber past a week."),
    ]
    for key, value, fmt, meaning in after:
        ref, label = DASH[key]
        row = ref[1:]
        put(ws, "A{0}".format(row), label, bold=True)
        put(ws, ref, formula(value), bold=True, size=14, fmt=fmt, color=PULLED_GREEN)
        note(ws, "C{0}".format(row), meaning)
        ws.merge_cells("C{0}:E{0}".format(row))
        ws.row_dimensions[int(row)].height = 24

    ws.conditional_formatting.add(DASH["days"][0], CellIsRule(
        operator="greaterThan", formula=["14"],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))
    ws.conditional_formatting.add(DASH["days"][0], CellIsRule(
        operator="greaterThan", formula=["7"],
        fill=PatternFill("solid", bgColor=AMBER_FILL)))
    ws.conditional_formatting.add(DASH["ale_pct"][0], CellIsRule(
        operator="greaterThan", formula=["0.8"],
        fill=PatternFill("solid", bgColor=AMBER_FILL)))

    put(ws, "G29", "Contractors", bold=True, size=11, color="FFFFFF", fill=BAR)
    ws.merge_cells("G29:J29")
    for i, text in enumerate(["Contractor", "Score", "Red flags", "Verdict"]):
        cell = ws.cell(row=30, column=7 + i, value=text)
        style(cell, bold=True, fill=HEAD, wrap=True, align="left")
        cell.border = Border(bottom=RULE)
    for i, col in enumerate(CONTRACTOR_COLS):
        row = DASH_CONTRACTOR_FIRST + i
        put(ws, "G{0}".format(row), formula(
            '=IF({0}!{1}${2}="","Not entered",{0}!{1}${2})'.format(
                CON, col, ROW_COMPANY)), color=PULLED_GREEN)
        put(ws, "H{0}".format(row), formula(
            '=IF({0}!{1}${2}="","",{0}!{1}${3})'.format(
                CON, col, ROW_COMPANY, ROW_SCORE)), align="center")
        put(ws, "I{0}".format(row), formula(
            '=IF({0}!{1}${2}="","",{0}!{1}${3})'.format(
                CON, col, ROW_COMPANY, ROW_FLAGS)), align="center")
        put(ws, "J{0}".format(row), formula(
            '=IF({0}!{1}${2}="","",{0}!{1}${3})'.format(
                CON, col, ROW_COMPANY, ROW_VERDICT)), bold=True)
    ws.conditional_formatting.add("J{0}:J{1}".format(
        DASH_CONTRACTOR_FIRST, DASH_CONTRACTOR_FIRST + len(CONTRACTOR_COLS) - 1), CellIsRule(
        operator="equal", formula=['"STOP, walk away"'],
        fill=PatternFill("solid", bgColor=RED_FILL), font=Font(bold=True, color="9C0006")))

    wrong = []
    for key, (ref, label) in DASH.items():
        got = ws["A{0}".format(ref[1:])].value
        if got != label:
            wrong.append("{0} -> {1}: label is {2!r}, DASH says {3!r}".format(key, ref, got, label))
    if wrong:
        raise SystemExit("the Dashboard moved under DASH:\n  " + "\n  ".join(wrong))

    widths(ws, {"A": 38, "B": 18, "C": 22, "D": 16, "E": 30, "F": 3, "G": 30,
                "H": 14, "I": 12, "J": 26})
    printing(ws, "A1:J{0}".format(
        DASH_CONTRACTOR_FIRST + len(CONTRACTOR_COLS) - 1), landscape=True)
    lock(ws)
    return ws


# ------------------------------------------------------------------- START HERE

START = [
    ("h", "The Gulf Coast Storm Season Workbook"),
    ("warn", "Store account and policy numbers as the last four digits only. Do not "
             "publish this file to the web and do not share it with an \"anyone with "
             "the link\" setting. If you keep a cloud copy, keep it in the same "
             "account you would keep a passport scan in. " + SECURITY_NOTE),
    ("p", "This is the companion to the Gulf Coast Storm Season Binder. The binder is "
          "what you carry: the countdown, the shutdown sequence, the re-entry walk, "
          "readable by a wet person with no power. This is what you calculate. It does "
          "the arithmetic paper cannot: deductibles in dollars, an inventory totaled "
          "against your contents limit, damage subtotaled by which policy pays, and "
          "receipts measured against a loss of use limit."),
    ("h2", "Do these four things, in this order"),
    ("n", "Fill in Coverage & Deductibles with your declarations page in front of you. "
          "It takes fifteen minutes and it is the only tab that matters before a "
          "storm is named. Most of this workbook is downstream of it."),
    ("n", "Read the Dashboard. If the gap figure is red, that is the number to work on "
          "between now and next season."),
    ("n", "Work Home Inventory one room at a time. Do not try to finish the house in "
          "one sitting. Start with the expensive things."),
    ("n", "Fill in the Supply Calculator, print the Still need column, and take it "
          "shopping before the cone is on the television."),
    ("h2", "The rest of the tabs are for after"),
    ("d", "Damage Log", "Subtotals what broke by cause, and tells you whether each "
                        "side is above its own deductible yet. Wind and flood are "
                        "separate policies and separate companies."),
    ("d", "Receipts & Loss of Use", "The page the binder does not have. Loss of use "
                                    "draws on its own limit. Mitigation spending does "
                                    "not. Keeping them apart is worth real money."),
    ("d", "Claim Call Log", "Get a name every time. A claim is decided by a file, and "
                            "the person keeping the better file usually wins the "
                            "argument about what was agreed six weeks ago."),
    ("d", "Contractor Comparison", "Ten checks, scored. Any single red flag overrides "
                                   "a perfect score, because one is enough."),
    ("h2", "Nothing about your policy is pre-filled, and that is deliberate"),
    ("p", "Every coverage limit, deductible and sub-limit cell starts empty. A "
          "plausible-looking default would produce a confident wrong answer on exactly "
          "the calculation you most need to be right, and you would never know. Where "
          "a cell is empty the workbook says so out loud rather than treating it as a "
          "zero. An unknown sub-limit is not the same thing as a sub-limit you are "
          "inside of, and the Dashboard will keep saying so until you look it up."),
    ("h2", "What the colors mean"),
    ("c", "Blue text", "Yours. Type over it."),
    ("c", "Black text", "A calculation. It will come back if you delete it, but leave "
                        "it alone."),
    ("c", "Green text", "Pulled in from another tab."),
    ("c", "Yellow fill", "An assumption, or an answer still missing. Review it."),
    ("c", "Gray fill", "Locked, or a state the workbook does not know yet."),
    ("h2", "The sheets are protected, and the password is nothing"),
    ("p", "Formula columns are locked so a stray paste cannot quietly break the math. "
          "There is no password: Review, Unprotect Sheet, and it opens. In Google "
          "Sheets you get a warning you can dismiss. It is a seatbelt, not a lock."),
    ("h2", "Every tab prints"),
    ("p", "Letter, fit to one page wide, with the header row repeating. Print the "
          "filled inventory and three-hole-punch it into the binder, because after "
          "landfall there is no laptop and no power. That is the whole reason the two "
          "products are shaped the way they are."),
    ("p", DISCLAIMER),
]


def sheet_start(wb):
    ws = wb.create_sheet("START HERE")
    row, step = 1, 0
    for item in START:
        kind, rest = item[0], item[1:]
        if kind == "h":
            put(ws, "A{0}".format(row), rest[0], bold=True, size=16, color=BAR)
            row += 2
        elif kind == "warn":
            cell = put(ws, "A{0}".format(row), rest[0], wrap=True, bold=True,
                       fill=YELLOW, color="7F4F00")
            cell.border = Border(bottom=RULE, top=RULE)
            ws.merge_cells("A{0}:C{1}".format(row, row + 2))
            ws.row_dimensions[row].height = 30
            row += 4
        elif kind == "h2":
            row += 1
            cell = put(ws, "A{0}".format(row), rest[0], bold=True, size=12, color=BAR)
            cell.border = Border(bottom=RULE)
            ws.merge_cells("A{0}:C{0}".format(row))
            row += 1
        elif kind == "p":
            put(ws, "A{0}".format(row), rest[0], wrap=True)
            ws.merge_cells("A{0}:C{1}".format(row, row + 2))
            ws.row_dimensions[row].height = 28
            row += 3
        elif kind == "n":
            step += 1
            put(ws, "A{0}".format(row), "{0}.".format(step), bold=True, align="right")
            put(ws, "B{0}".format(row), rest[0], wrap=True)
            ws.merge_cells("B{0}:C{1}".format(row, row + 1))
            ws.row_dimensions[row].height = 26
            row += 2
        else:  # "d" and "c" are both a labeled definition
            put(ws, "A{0}".format(row), rest[0], bold=True,
                color=INPUT_BLUE if kind == "c" else INK)
            put(ws, "B{0}".format(row), rest[1], wrap=True)
            ws.merge_cells("B{0}:C{1}".format(row, row + 1))
            ws.row_dimensions[row].height = 26
            row += 2
    widths(ws, {"A": 24, "B": 58, "C": 40})
    printing(ws, "A1:C{0}".format(row))
    lock(ws)
    return ws


# ----------------------------------------------------------------------- build

ORDER = ["START HERE", "Dashboard", "Coverage & Deductibles", "Home Inventory",
         "Supply Calculator", "Damage Log", "Receipts & Loss of Use",
         "Claim Call Log", "Contractor Comparison", "Lists"]


def build(out=OUT):
    rooms = check_against_binder()

    wb = Workbook()
    wb.remove(wb.active)
    # Arial everywhere, including the cells nothing ever writes to. Left alone,
    # openpyxl stamps them Calibri, which Google Sheets substitutes on import and
    # which is a font we have no reason to be shipping. Index 0 is the record
    # every unstyled cell points at, so it has to be replaced as well as the
    # Normal style.
    base = Font(name="Arial", size=10, color=INK)
    wb._named_styles["Normal"].font = base
    wb._fonts[0] = base

    # Lists first, because every dropdown on every other tab is a named range
    # defined here. Coverage next, because everything downstream reads it.
    sheet_lists(wb)
    sheet_coverage(wb)
    sheet_inventory(wb)
    sheet_supply(wb)
    sheet_damage(wb)
    sheet_receipts(wb)
    sheet_calls(wb)
    sheet_contractors(wb)
    sheet_dashboard(wb)
    sheet_start(wb)

    wb._sheets = [wb[name] for name in ORDER]
    wb.active = 0

    folder = os.path.dirname(out)
    if folder and not os.path.isdir(folder):
        os.makedirs(folder)
    wb.save(out)

    print("{0}: {1} tabs, {2} rooms locked to the binder, {3} inventory rows, "
          "{4} sub-limits tracked".format(
              out, len(ORDER), rooms, INV_LAST - INV_FIRST + 1, len(SPECIAL)))
    return out


if __name__ == "__main__":
    sys.exit(0 if build() else 1)
