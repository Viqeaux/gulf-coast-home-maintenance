"""Runs the storm workbook's QA cases against a real formula engine.

    python qa_workbook.py
    python qa_workbook.py --checklist

`build_workbook.py` writes formulas; it cannot evaluate them, and openpyxl never
will. So this builds a small copy of the same workbook, compiles it with the
`formulas` package, and reads the answers back out. Every number below is
computed, not eyeballed.

That distinction is the whole point of the file. An off-by-one reference
produces a workbook with no visible errors in it and quietly wrong numbers,
which is worse than a clean break: the buyer never finds out, and neither do we.
On this product the stakes are higher than on the planner, because the number
being got wrong is what a hurricane costs somebody before their policy pays.

**Small logs, not full ones.** `build_workbook.configure()` shrinks the row
blocks and nothing else. Every formula, every reference and every summary cell
is the one that ships, so this is a test of the product rather than of a model
of it. Five hundred inventory rows is more than a pure-Python engine will chew
through in a sitting.

**What this cannot reach.** The engine is not Excel and is not Google Sheets, so
it says nothing about whether the dropdowns, the conditional formatting, the
print setup or the protection survived. Those are file structure, and
`structure_checks()` reads them straight out of the shipped .xlsx with openpyxl.
What is left after that is a short list a human has to look at once, in Sheets,
after importing: `--checklist` prints it.

Twelve cases, from the build spec:

    1   an empty workbook shows no errors anywhere
    2   $400,000 at 2 percent resolves to exactly $8,000
    3   toggling to a flat dollar amount switches D6 cleanly
    4   a blank contents limit falls back to the percentage of dwelling
    5   the under-listing nudge fires all four of its states
    6   a sub-limit flag fires when the category total passes the limit
    6a  a blank sub-limit with items owned never reads "Within limit"
    6b  a blank sub-limit with nothing owned stays quiet and gray
    6c  a fresh copy counts 6 of 6 sub-limits not entered
    7   4 people and 7 days gives 28 gallons, 84 meals and $400
    8   the generator set to No returns zero fuel
    9   wind side and flood side move independently, Unknown lands in neither
    10  mitigation spending does not count toward the loss of use limit
    11  one red flag with a perfect score still returns STOP
    12  the file structure survived the write
"""

import os
import re
import sys
import tempfile

from openpyxl import load_workbook

import build_workbook as bw
from workbook_data import SPECIAL

QA_INVENTORY = 8
QA_LOGS = 6
QA_CALLS = 6

COVERAGE = "COVERAGE & DEDUCTIBLES"
INVENTORY = "HOME INVENTORY"
SUPPLY = "SUPPLY CALCULATOR"
DAMAGE = "DAMAGE LOG"
RECEIPTS = "RECEIPTS & LOSS OF USE"
CALLS = "CLAIM CALL LOG"
CONTRACTORS = "CONTRACTOR COMPARISON"
DASHBOARD = "DASHBOARD"

CELL = re.compile(r"^'\[[^\]]+\](?P<sheet>[^']+)'!(?P<ref>\$?[A-Z]{1,3}\$?\d+)$")
HIDDEN_FROM_BUYER = {"LISTS"}

# Which Dashboard exposure row each special category landed on.
EXPOSURE = {name: bw.DASH_SUBLIMIT_FIRST + i for i, name in enumerate(SPECIAL)}
SUBLIMIT = {name: bw.SUBLIMIT_FIRST + i for i, name in enumerate(SPECIAL)}


class Model(object):
    """The compiled workbook, plus whatever the current case has typed into it."""

    def __init__(self, path):
        import formulas
        self.book = os.path.basename(path)
        self.model = formulas.ExcelModel().loads(path).finish()
        self.inputs = {}
        self.solution = None

    def key(self, sheet, ref):
        return "'[{0}]{1}'!{2}".format(self.book, sheet, ref)

    def reset(self):
        self.inputs = {}
        self.solution = None

    def set(self, sheet, ref, value):
        self.inputs[self.key(sheet, ref)] = value
        self.solution = None

    def clear(self, sheet, ref):
        self.set(sheet, ref, "")

    def solve(self):
        self.solution = self.model.calculate(inputs=dict(self.inputs))
        return self.solution

    def get(self, sheet, ref):
        if self.solution is None:
            self.solve()
        node = self.solution.get(self.key(sheet, ref))
        if node is None:
            raise KeyError("{0}!{1} is not in the model".format(sheet, ref))
        value = node.value[0, 0]
        return value.tolist() if hasattr(value, "tolist") else value

    def errors(self):
        """Every visible cell currently showing an Excel error."""
        if self.solution is None:
            self.solve()
        found = []
        for key, node in self.solution.items():
            match = CELL.match(key)
            if not match or match.group("sheet").upper() in HIDDEN_FROM_BUYER:
                continue
            try:
                value = node.value[0, 0]
            except Exception:
                continue
            if isinstance(value, str) and value.startswith("#"):
                found.append("{0}!{1} {2}".format(match.group("sheet"), match.group("ref"), value))
        return sorted(set(found))


class Case(object):
    def __init__(self, model, name):
        self.m = model
        self.name = name
        self.failures = []
        model.reset()

    # --- typing into the workbook

    def policy(self, dwelling=None, wind_type=None, wind_pct=None, wind_flat=None,
               contents=None, contents_pct=None, lou=None, flood_ded=None,
               savings=None):
        pairs = [("B4", dwelling), ("B6", wind_type), ("B7", wind_pct),
                 ("B8", wind_flat), ("B9", contents), ("B10", contents_pct),
                 ("B11", lou), ("B15", flood_ded), ("B16", savings)]
        for ref, value in pairs:
            if value is not None:
                self.m.set(COVERAGE, ref, value)

    def sublimit(self, category, value):
        self.m.set(COVERAGE, "B{0}".format(SUBLIMIT[category]), value)

    def item(self, offset, room=None, name="Thing", qty=1, cost=0, category=None,
             photographed=None):
        row = bw.INV_FIRST + offset
        self.m.set(INVENTORY, "B{0}".format(row), name)
        self.m.set(INVENTORY, "F{0}".format(row), qty)
        self.m.set(INVENTORY, "G{0}".format(row), cost)
        if room is not None:
            self.m.set(INVENTORY, "A{0}".format(row), room)
        if category is not None:
            self.m.set(INVENTORY, "I{0}".format(row), category)
        if photographed is not None:
            self.m.set(INVENTORY, "J{0}".format(row), photographed)

    def damage(self, offset, cause, value):
        row = bw.DMG_FIRST + offset
        self.m.set(DAMAGE, "C{0}".format(row), "Something")
        self.m.set(DAMAGE, "D{0}".format(row), cause)
        self.m.set(DAMAGE, "F{0}".format(row), value)

    def receipt(self, offset, category, amount, submitted=None, reimbursed=None):
        row = bw.RCP_FIRST + offset
        self.m.set(RECEIPTS, "B{0}".format(row), category)
        self.m.set(RECEIPTS, "D{0}".format(row), amount)
        if submitted is not None:
            self.m.set(RECEIPTS, "G{0}".format(row), submitted)
        if reimbursed is not None:
            self.m.set(RECEIPTS, "I{0}".format(row), reimbursed)

    def supply(self, people=None, pets=None, days=None, coolers=None,
               vehicles=None, generator=None, hours=None):
        pairs = [("B4", people), ("B5", pets), ("B6", days), ("B7", coolers),
                 ("B8", vehicles), ("B9", generator), ("B10", hours)]
        for ref, value in pairs:
            if value is not None:
                self.m.set(SUPPLY, ref, value)

    # --- reading it back

    def cov(self, ref):
        return self.m.get(COVERAGE, ref)

    def dash(self, key):
        return self.m.get(DASHBOARD, bw.DASH[key][0])

    def exposure(self, category, col):
        return self.m.get(DASHBOARD, "{0}{1}".format(col, EXPOSURE[category]))

    def supply_row(self, label, col="C"):
        for i, (name, _rule, _unit, _value, _fmt) in enumerate(bw.supply_rows()):
            if name == label:
                return self.m.get(SUPPLY, "{0}{1}".format(col, bw.SUP_FIRST + i))
        raise KeyError(label)

    # --- assertions

    def check(self, label, got, want, tolerance=0.51):
        if isinstance(want, (int, float)) and not isinstance(want, bool):
            try:
                ok = abs(float(got) - float(want)) <= tolerance
            except (TypeError, ValueError):
                ok = False
        else:
            ok = got == want
        if not ok:
            self.failures.append("{0}: got {1!r}, wanted {2!r}".format(label, got, want))
        return ok

    def starts(self, label, got, prefix):
        ok = isinstance(got, str) and got.startswith(prefix)
        if not ok:
            self.failures.append("{0}: got {1!r}, wanted it to start {2!r}".format(
                label, got, prefix))
        return ok

    def assert_true(self, label, condition):
        if not condition:
            self.failures.append(label)
        return condition

    def is_blank(self, label, value):
        return self.assert_true("{0}: got {1!r}".format(label, value), value in (None, ""))

    def no_errors(self, label="no error cells"):
        found = self.m.errors()
        return self.assert_true("{0}, found {1}".format(label, found[:6]), not found)


# ------------------------------------------------------------------------ cases

def case_1_empty(case):
    """A buyer who has opened the file and typed nothing sees no breakage."""
    case.no_errors()

    # The wind deductible refuses to answer rather than falling through to the
    # flat-amount branch and printing a confident $0. This is the single most
    # dangerous wrong number the workbook could produce.
    case.is_blank("wind deductible with no type chosen", case.cov("D6"))
    case.is_blank("combined worst case", case.cov("D17"))
    case.is_blank("the gap", case.cov("D18"))

    case.check("contents listed", case.dash("listed"), 0)
    case.check("share of limit", case.dash("pct"), 0)
    case.check("items logged", case.dash("items"), 0)
    case.check("photographed", case.dash("photographed"), 0)
    case.check("wind side damage", case.dash("wind_damage"), 0)
    case.check("loss of use share", case.dash("ale_pct"), 0)
    case.is_blank("days since last call", case.dash("days"))

    # The tabs downstream say which tab to go to rather than comparing against
    # a blank and reporting "below your deductible, no payout", which would be
    # a reassuring sentence built on a number nobody has entered.
    case.check("damage log, wind verdict", case.m.get(DAMAGE, "C6"), bw.UNSET)
    case.check("damage log, flood verdict", case.m.get(DAMAGE, "C7"), bw.UNSET)


def case_2_percentage(case):
    """$400,000 at 2 percent is $8,000. This is the listing screenshot."""
    case.policy(dwelling=400000, wind_type="Percentage of dwelling", wind_pct=0.02,
                flood_ded=0, savings=0)
    case.check("wind deductible", case.cov("D6"), 8000)
    case.check("dashboard agrees", case.dash("wind"), 8000)

    case.policy(flood_ded=2000, savings=3000)
    case.check("both deductibles", case.cov("D17"), 10000)
    case.check("what you would have to find", case.cov("D18"), 7000)

    case.policy(savings=50000)
    case.check("savings above the worst case closes the gap", case.cov("D18"), 0)
    case.no_errors()


def case_3_flat(case):
    """Toggling the type switches D6 cleanly, with no stale percentage left in."""
    case.policy(dwelling=400000, wind_type="Percentage of dwelling", wind_pct=0.02,
                wind_flat=1500, flood_ded=0, savings=0)
    case.check("percentage branch", case.cov("D6"), 8000)
    case.policy(wind_type="Flat dollar amount")
    case.check("flat branch", case.cov("D6"), 1500)
    case.check("the percentage did not leak through", case.cov("D17"), 1500)
    case.no_errors()


def case_4_contents_fallback(case):
    """A blank contents limit falls back to the share of dwelling, not to zero."""
    case.policy(dwelling=400000, contents_pct=0.5)
    case.m.clear(COVERAGE, "B9")
    case.check("fallback used", case.cov("D9"), 200000)
    case.check("dashboard agrees", case.dash("contents_limit"), 200000)

    case.policy(contents=125000)
    case.check("a real figure beats the fallback", case.cov("D9"), 125000)
    case.no_errors()


def case_5_nudge(case):
    """All four states of the under-listing nudge, as the value crosses the lines."""
    nudge = lambda: case.m.get(INVENTORY, "A3")

    case.policy(dwelling=400000, contents_pct=0.5)
    case.m.clear(COVERAGE, "B9")
    case.starts("empty inventory", nudge(), "Start with the expensive things")

    case.item(0, room=None, name="Sofa", qty=1, cost=1000)
    case.starts("well under half", nudge(), "You have listed less than half")

    case.item(1, name="Everything else", qty=1, cost=150000)
    case.starts("comfortably inside", nudge(), "Good coverage of your limit")

    case.item(2, name="And the rest", qty=1, cost=100000)
    case.starts("over the limit", nudge(), "You have listed more than your contents limit")
    case.check("share of limit", case.dash("pct"), 1.255, tolerance=0.01)
    case.no_errors()


def case_6_sublimit_over(case):
    """Jewelry past its limit reads OVER, with the exposure in dollars."""
    case.sublimit("Jewelry", 2500)
    case.item(0, name="Rings", qty=1, cost=9000, category="Jewelry")
    case.check("value owned", case.exposure("Jewelry", "B"), 9000)
    case.check("limit pulled through", case.exposure("Jewelry", "C"), 2500)
    case.check("exposure", case.exposure("Jewelry", "D"), 6500)
    case.starts("flag", case.exposure("Jewelry", "E"), "OVER SUB-LIMIT")

    # Under the limit is the quiet, correct green state.
    case.sublimit("Jewelry", 15000)
    case.check("within limit", case.exposure("Jewelry", "E"), "Within limit")
    case.check("exposure closes", case.exposure("Jewelry", "D"), 0)
    case.no_errors()


def case_6a_blank_with_value(case):
    """The one that matters. A blank limit must never read "Within limit".

    `='Coverage & Deductibles'!$B$19` on an empty cell is 0, not blank, and a
    naive comparison then reports the buyer as fine on a category they have not
    looked up. The exposure column stays empty rather than showing 0, because a
    zero there reads as "nothing to worry about".
    """
    case.item(0, name="Grandmother's rings", qty=1, cost=9000, category="Jewelry")
    case.check("value owned is seen", case.exposure("Jewelry", "B"), 9000)
    case.is_blank("limit stays blank across the hop", case.exposure("Jewelry", "C"))
    case.is_blank("exposure stays blank", case.exposure("Jewelry", "D"))

    flag = case.exposure("Jewelry", "E")
    case.starts("flag escalates", flag, "LIMIT NOT ENTERED")
    case.assert_true("flag names the dollar figure, got {0!r}".format(flag),
                     "9,000" in str(flag))
    case.assert_true("flag never says Within limit, got {0!r}".format(flag),
                     "Within limit" not in str(flag))
    case.no_errors()


def case_6b_blank_with_nothing(case):
    """Nothing owned in the category and no limit entered is not an alarm."""
    case.check("quiet state", case.exposure("Firearms", "E"), "Limit not entered")
    case.is_blank("no exposure", case.exposure("Firearms", "D"))
    case.check("nothing owned", case.exposure("Firearms", "B"), 0)
    case.no_errors()


def case_6c_fresh_count(case):
    """A fresh copy reads 6 of 6, and the count falls as they are entered."""
    case.check("fresh copy", case.dash("unentered"), len(SPECIAL))
    case.sublimit("Jewelry", 2500)
    case.check("one entered", case.dash("unentered"), len(SPECIAL) - 1)
    case.sublimit("Cash", 200)
    case.check("two entered", case.dash("unentered"), len(SPECIAL) - 2)

    # A category the buyer owns things in but has not looked up still counts.
    case.item(0, name="Shotgun", qty=1, cost=1200, category="Firearms")
    case.check("owning something does not clear it", case.dash("unentered"),
               len(SPECIAL) - 2)
    case.no_errors()


def case_7_supply(case):
    """Four people, seven days. 28 gallons, 84 meals, $400."""
    case.supply(people=4, pets=1, days=7, coolers=2, vehicles=2,
                generator="Yes", hours=9)
    case.check("drinking water", case.supply_row("Drinking water"), 28)
    case.check("cooking and washing", case.supply_row("Water for cooking and washing"), 14)
    case.check("pet water", case.supply_row("Pet water"), 7)
    case.check("meals", case.supply_row("Non perishable food"), 84)
    case.check("ice", case.supply_row("Ice"), 14)
    case.check("cash", case.supply_row("Cash, in small bills"), 400)
    case.check("vehicle tanks", case.supply_row("Vehicle fuel"), 2)
    case.check("generator fuel", case.supply_row("Generator fuel"), 35)

    # Still need is the shopping list, and it never goes negative.
    case.m.set(SUPPLY, "D{0}".format(bw.SUP_FIRST), 10)
    case.check("still need", case.supply_row("Drinking water", col="E"), 18)
    case.m.set(SUPPLY, "D{0}".format(bw.SUP_FIRST), 40)
    case.check("having more than enough is zero, not negative",
               case.supply_row("Drinking water", col="E"), 0)
    case.no_errors()


def case_8_no_generator(case):
    case.supply(people=4, days=7, generator="No", hours=9)
    case.check("no generator, no fuel", case.supply_row("Generator fuel"), 0)
    case.check("propane is two either way", case.supply_row("Propane"), 2)
    case.supply(generator="Yes")
    case.check("and it comes back", case.supply_row("Generator fuel"), 35)
    case.no_errors()


def case_9_damage_sides(case):
    """The two sides move independently and Unknown lands in neither."""
    case.policy(dwelling=400000, wind_type="Percentage of dwelling", wind_pct=0.02,
                flood_ded=2000)
    case.damage(0, "Wind", 3000)
    case.damage(1, "Wind-driven rain", 1500)
    case.damage(2, "Tree", 500)
    case.damage(3, "Flood", 4000)
    case.damage(4, "Surge", 1000)
    case.damage(5, bw.UNCATEGORIZED, 7000)

    case.check("wind side", case.dash("wind_damage"), 5000)
    case.check("flood side", case.dash("flood_damage"), 5000)
    case.check("uncategorized", case.dash("uncat"), 7000)
    case.assert_true("Unknown stayed out of the wind side",
                     case.dash("wind_damage") == 5000)

    # Wind-driven rain must not be swept into a "Wind*" match. 5,000 rather
    # than 6,500 is the whole test.
    case.check("wind verdict", case.m.get(DAMAGE, "C6"),
               "Below your wind deductible, currently no payout")
    case.check("flood verdict", case.m.get(DAMAGE, "C7"),
               "Above your flood deductible")

    case.damage(0, "Wind", 10000)
    case.check("wind side moved alone", case.dash("wind_damage"), 12000)
    case.check("flood side did not", case.dash("flood_damage"), 5000)
    case.check("wind verdict flipped", case.m.get(DAMAGE, "C6"),
               "Above your wind deductible")
    case.no_errors()


def case_10_mitigation(case):
    """Mitigation is reimbursable and is not drawn from the loss of use limit."""
    case.policy(dwelling=400000, lou=80000)
    case.receipt(0, "Lodging", 4000, submitted="Yes", reimbursed=1000)
    case.receipt(1, "Meals", 800, submitted="Yes")
    case.receipt(2, "Mitigation supplies", 3000)
    case.receipt(3, "Generator fuel", 200)
    case.receipt(4, "Other", 500)

    case.check("loss of use subtotal", case.m.get(RECEIPTS, "B5"), 4800)
    case.check("mitigation subtotal", case.m.get(RECEIPTS, "E6"), 3200)
    case.check("loss of use share of limit", case.m.get(RECEIPTS, "B6"), 0.06,
               tolerance=0.001)
    case.check("total spent", case.m.get(RECEIPTS, "B7"), 8500)
    case.check("total submitted", case.m.get(RECEIPTS, "E7"), 4800)
    case.check("total reimbursed", case.m.get(RECEIPTS, "B8"), 1000)
    case.check("outstanding", case.m.get(RECEIPTS, "E8"), 7500)

    # Pile mitigation on and the loss of use percentage must not move an inch.
    case.receipt(2, "Mitigation supplies", 40000)
    case.check("loss of use share held", case.m.get(RECEIPTS, "B6"), 0.06,
               tolerance=0.001)
    case.check("dashboard agrees", case.dash("ale_pct"), 0.06, tolerance=0.001)
    case.check("mitigation is still counted somewhere", case.dash("mitigation"), 40200)
    case.no_errors()


def case_11_contractor(case):
    """One red flag with a perfect score is still STOP. Any one is enough."""
    for row in range(bw.CHECK_FIRST, bw.CHECK_LAST + 1):
        case.m.set(CONTRACTORS, "C{0}".format(row), "Yes")
    case.m.set(CONTRACTORS, "C{0}".format(bw.ROW_COMPANY), "Someone Roofing")
    case.check("perfect score", case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_SCORE)), 10)
    case.check("cleared", case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_VERDICT)),
               "Cleared")

    case.m.set(CONTRACTORS, "C{0}".format(bw.FLAG_FIRST + 4), "Yes")
    case.check("one red flag", case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_FLAGS)), 1)
    case.check("a perfect score does not survive it",
               case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_VERDICT)), "STOP, walk away")
    case.check("dashboard carries the verdict",
               case.m.get(DASHBOARD, "J{0}".format(bw.DASH_CONTRACTOR_FIRST)),
               "STOP, walk away")

    # The middle states, so the thresholds are not off by one.
    case.m.set(CONTRACTORS, "C{0}".format(bw.FLAG_FIRST + 4), "No")
    case.m.set(CONTRACTORS, "C{0}".format(bw.CHECK_LAST), "Pending")
    case.check("nine of ten is cleared",
               case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_VERDICT)), "Cleared")
    case.m.set(CONTRACTORS, "C{0}".format(bw.CHECK_LAST - 1), "Pending")
    case.check("eight of ten is incomplete",
               case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_VERDICT)),
               "Incomplete, finish the checks")
    for row in range(bw.CHECK_FIRST, bw.CHECK_LAST + 1):
        case.m.set(CONTRACTORS, "C{0}".format(row), "Pending")
    case.check("nothing done is not vetted",
               case.m.get(CONTRACTORS, "C{0}".format(bw.ROW_VERDICT)), "Not vetted")
    case.no_errors()


CASES = [
    ("1   empty workbook", case_1_empty),
    ("2   percentage deductible", case_2_percentage),
    ("3   flat dollar toggle", case_3_flat),
    ("4   contents fallback", case_4_contents_fallback),
    ("5   under-listing nudge", case_5_nudge),
    ("6   sub-limit exceeded", case_6_sublimit_over),
    ("6a  blank limit, items owned", case_6a_blank_with_value),
    ("6b  blank limit, nothing owned", case_6b_blank_with_nothing),
    ("6c  fresh copy counts 6 of 6", case_6c_fresh_count),
    ("7   supply arithmetic", case_7_supply),
    ("8   generator set to No", case_8_no_generator),
    ("9   damage sides", case_9_damage_sides),
    ("10  mitigation vs loss of use", case_10_mitigation),
    ("11  contractor red flag", case_11_contractor),
]


# ------------------------------------------------------- case 12, file structure

def structure_checks(path):
    """What the formula engine cannot see, read out of the shipped file."""
    problems = []

    def want(label, got, expected):
        if got != expected:
            problems.append("{0}: got {1!r}, wanted {2!r}".format(label, got, expected))

    wb = load_workbook(path)
    want("tab order", wb.sheetnames, bw.ORDER)
    want("Lists is hidden", wb["Lists"].sheet_state, "hidden")
    for name in wb.sheetnames:
        if name != "Lists":
            want("{0} is visible".format(name), wb[name].sheet_state, "visible")

    names = set(wb.defined_names)
    for _title, listname, _values in bw.LIST_COLUMNS:
        if listname not in names:
            problems.append("named range {0} is missing".format(listname))

    # Dropdowns, per tab and per column. A lost validation is invisible until a
    # buyer types "wind driven rain" and a SUMIF silently stops matching it.
    expected_dv = {
        "Home Inventory": {"A": "=RoomList", "I": "=CategoryList", "J": "=YesNoList"},
        "Damage Log": {"D": "=CauseList", "G": "=RepairedList"},
        "Receipts & Loss of Use": {"B": "=ReceiptList", "G": "=YesNoList",
                                   "H": "=ReimbursedList"},
        "Claim Call Log": {"J": "=YesNoList"},
        "Contractor Comparison": {"C": "=CheckList"},
        "Supply Calculator": {"B": "=YesNoList"},
        "Coverage & Deductibles": {"B": "=WindTypeList"},
    }
    for sheet, columns in expected_dv.items():
        found = {}
        for dv in wb[sheet].data_validations.dataValidation:
            for cells in dv.sqref.ranges:
                found.setdefault(cells.coord.split(":")[0][0], []).append(dv.formula1)
        for col, listname in columns.items():
            if listname not in found.get(col, []):
                problems.append("{0} column {1}: no {2} dropdown, found {3}".format(
                    sheet, col, listname, found.get(col)))

    # The sub-limit cells ship empty, yellow and refusing a zero. All three are
    # load-bearing: a plausible default here produces a confident wrong answer.
    cov = wb["Coverage & Deductibles"]
    for row in range(bw.SUBLIMIT_FIRST, bw.SUBLIMIT_LAST + 1):
        cell = cov["B{0}".format(row)]
        if cell.value is not None:
            problems.append("sub-limit B{0} ships with {1!r} in it".format(row, cell.value))
        if cell.protection.locked:
            problems.append("sub-limit B{0} is locked".format(row))
        if not (cell.fill and cell.fill.fgColor.rgb and
                cell.fill.fgColor.rgb.endswith(bw.YELLOW)):
            problems.append("sub-limit B{0} is not yellow-filled".format(row))
        if not cov["C{0}".format(row)].value:
            problems.append("sub-limit B{0} has no helper text beside it".format(row))
    positive = [dv for dv in cov.data_validations.dataValidation
                if dv.type == "decimal" and dv.operator == "greaterThan"]
    want("sub-limits refuse a zero", len(positive), 1)

    # No policy figure ships pre-filled. Only the two share-of-dwelling
    # fallbacks and the two supply assumptions carry a default anywhere.
    for row, _label, default, _list, _kind, _fmt, _why in bw.COVERAGE_INPUTS:
        got = cov["B{0}".format(row)].value
        if row in (10, 12):
            want("fallback B{0} keeps its default".format(row), got, default)
        elif got is not None:
            problems.append("B{0} ships pre-filled with {1!r}".format(row, got))

    inv = wb["Home Inventory"]
    for col in "ABCDEFGIJKL":
        if inv.column_dimensions[col].protection.locked:
            problems.append("Home Inventory column {0} is locked but is the buyer's"
                            .format(col))
    if not inv["H{0}".format(bw.INV_FIRST)].protection.locked:
        problems.append("the line total column is unlocked")
    filled = sum(1 for r in range(bw.INV_FIRST, bw.INV_LAST + 1)
                 if inv["H{0}".format(r)].value)
    want("line total formulas", filled, bw.INV_LAST - bw.INV_FIRST + 1)
    for col in "ABCDEFG":
        if inv["{0}{1}".format(col, bw.INV_FIRST)].value is not None:
            problems.append("Home Inventory column {0} is not empty for the buyer"
                            .format(col))

    for name in wb.sheetnames:
        ws = wb[name]
        if not ws.protection.sheet:
            problems.append("{0} is not protected".format(name))
        if ws.protection.password:
            problems.append("{0} is password protected, which it must never be".format(name))
        # Every tab the buyer can see prints, or the pair with the paper binder
        # does not work. Lists is hidden and has nothing to print.
        if ws.sheet_state != "visible":
            continue
        if not ws.print_area:
            problems.append("{0} has no print area".format(name))
        elif ws.page_setup.fitToWidth != 1:
            problems.append("{0} is not fit to one page wide".format(name))
    for name, orientation in (("Home Inventory", "landscape"), ("Damage Log", "landscape"),
                              ("Receipts & Loss of Use", "landscape"),
                              ("Claim Call Log", "landscape"),
                              ("Coverage & Deductibles", "portrait")):
        want("{0} orientation".format(name), wb[name].page_setup.orientation, orientation)
    for name, first in (("Home Inventory", bw.INV_HEAD), ("Damage Log", bw.DMG_HEAD),
                        ("Receipts & Loss of Use", bw.RCP_HEAD),
                        ("Claim Call Log", bw.CALL_HEAD)):
        want("{0} repeats its header row".format(name),
             str(wb[name].print_title_rows).replace("$", ""), "{0}:{0}".format(first))

    dash = wb["Dashboard"]
    want("Dashboard conditional formats", len(list(dash.conditional_formatting)) > 0, True)
    for key, (ref, label) in bw.DASH.items():
        if dash["A{0}".format(ref[1:])].value != label:
            problems.append("Dashboard {0} label moved".format(key))

    fonts = set()
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.font and cell.font.name:
                    fonts.add(cell.font.name)
    if fonts - {"Arial"}:
        problems.append("non-default fonts in the file: {0}".format(sorted(fonts - {"Arial"})))

    return problems


CHECKLIST = """
Six things no automation here can reach. Do them once, after importing the
.xlsx into Google Sheets.

  1. File, Import, Replace spreadsheet. Not "Open with": that leaves it as an
     uploaded file rather than a Sheet, and half of this list will look wrong.
  2. Home Inventory column A shows the nine rooms, and I and J show their own
     lists. Then Damage Log column D, which is the one a subtotal depends on.
  3. Coverage & Deductibles B19 to B24 are empty and yellow, and typing 0 into
     one of them is refused. A zero and "I have not looked it up" are different
     answers and only one of them is safe.
  4. The Dashboard exposure table is colored: red over the limit, amber for a
     category you own things in with no limit entered, gray for one you own
     nothing in, green for within. Four states, not three.
  5. Print preview every tab at Letter. Inventory, Damage Log, Receipts and
     Call Log should be landscape, one page wide, with the header row repeating
     on page two. This is what three-hole-punches into the binder.
  6. Every tab is still protected, Lists is still hidden, and the warning can be
     dismissed without a password.

The listing screenshot is Coverage & Deductibles with a $400,000 dwelling limit
and a 2 percent wind deductible, showing $8,000, with the gap below it in red.
Take it after step 1, at 100 percent zoom.

Then share as view only, and hand out the /copy link rather than the /edit one.
Never share edit access to the master.
""".strip()


def main():
    if "--checklist" in sys.argv:
        print(CHECKLIST)
        return 0

    print("building the shipping workbook")
    shipped = bw.build()
    problems = structure_checks(shipped)
    if problems:
        print("FAIL  12  file structure")
        for line in problems:
            print("          {0}".format(line))
    else:
        print("ok    12  file structure")

    print("building a smaller copy to calculate")
    bw.configure(inventory=QA_INVENTORY, logs=QA_LOGS, calls=QA_CALLS)
    path = os.path.join(tempfile.gettempdir(), "qa-storm-workbook.xlsx")
    bw.build(path)
    model = Model(path)

    failed = 1 if problems else 0
    for name, run in CASES:
        case = Case(model, name)
        try:
            run(case)
        except Exception as exc:
            case.failures.append("crashed: {0!r}".format(exc))
        if case.failures:
            failed += 1
            print("FAIL  {0}".format(name))
            for line in case.failures:
                print("          {0}".format(line))
        else:
            print("ok    {0}".format(name))

    print()
    if failed:
        print("{0} of {1} cases failed".format(failed, len(CASES) + 1))
    else:
        print("all {0} cases pass. Now run --checklist for the Google Sheets half."
              .format(len(CASES) + 1))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
