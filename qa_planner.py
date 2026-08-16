"""Runs the planner's QA cases against a real formula engine.

    python qa_planner.py
    python qa_planner.py --checklist

`build_planner.py` writes formulas; it cannot evaluate them, and openpyxl never
will. So this builds a six-row copy of the same workbook, compiles it with the
`formulas` package, and reads the answers back out. Every number below is
computed, not eyeballed.

That distinction is the whole point. An off-by-one reference produces a
workbook with no visible errors in it and quietly wrong numbers, which is worse
than a clean break: the buyer never finds out, and neither do we.

**Six rows, not a hundred.** `build_planner.configure()` shrinks the register
and nothing else. Every formula, every reference and every summary row is the
one that ships, so this is a test of the product rather than of a model of it.
A hundred rows times thirty years is 3,000 formulas and more than a pure-Python
engine will chew through in a sitting.

**What this cannot reach.** The engine is not Excel and is not Google Sheets,
so it says nothing about whether the dropdowns, the chart, the conditional
formatting or the protection survived. Those are file structure, and
`structure_checks()` reads them straight out of the shipped .xlsx with openpyxl.
What is left after that is five things a human has to look at once, in Sheets,
after importing: `--checklist` prints them.

Eleven cases. One to ten are the build spec's; eleven arrived with the Quick
Check tab:

    1  an empty workbook shows no errors anywhere
    2  an overdue item reads Overdue and lands in the current year's forecast
    3  a repeating item fires every cycle inside the window
    4  IDK - Assume On Schedule dates from the last cycle, not the build year
    5  an override beats the default in every downstream column
    6  switching region moves the costs and the exposure-driven lifespans
    7  zero inflation makes future cost equal today's cost exactly
    8  a thin contribution turns the reserve balance red and raises SHORTFALL
    9  two of the same system in different places track independently
    10 dropdowns, the hidden tab and the protection survived the write
    11 Quick Check answers from the build year alone, and its headline counts
"""

import math
import os
import re
import sys
import tempfile

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

import build_planner as bp

QA_ROWS = 6
EXAMPLE = bp.FIRST            # register row the example lives on
SECOND = bp.FIRST + 1
EXAMPLE_FC = bp.FC_FIRST      # the forecast row that mirrors the example
SECOND_FC = bp.FC_FIRST + 1

REGISTER = "SYSTEMS REGISTER"
SETUP = "SETUP"
FORECAST = "30-YEAR FORECAST"
DASHBOARD = "DASHBOARD"
QUICK = "QUICK CHECK"

CELL = re.compile(r"^'\[[^\]]+\](?P<sheet>[^']+)'!(?P<ref>\$?[A-Z]{1,3}\$?\d+)$")
HIDDEN_FROM_BUYER = {"LISTS"}


class Model(object):
    """The compiled workbook, plus whatever the current case has typed into it."""

    def __init__(self, path):
        import formulas
        self.book = os.path.basename(path)
        self.model = formulas.ExcelModel().loads(path).finish()
        self.inputs = {}
        self.solution = None

    def key(self, sheet, ref):
        return "'[{0}]{1}'!{2}".format(self.book, sheet, ref)

    def reset(self):
        self.inputs = {}
        self.solution = None

    def set(self, sheet, ref, value):
        self.inputs[self.key(sheet, ref)] = value
        self.solution = None

    def clear(self, sheet, ref):
        self.set(sheet, ref, "")

    def solve(self):
        self.solution = self.model.calculate(inputs=dict(self.inputs))
        return self.solution

    def get(self, sheet, ref):
        if self.solution is None:
            self.solve()
        node = self.solution.get(self.key(sheet, ref))
        if node is None:
            raise KeyError("{0}!{1} is not in the model".format(sheet, ref))
        value = node.value[0, 0]
        return value.tolist() if hasattr(value, "tolist") else value

    def errors(self):
        """Every visible cell currently showing an Excel error."""
        if self.solution is None:
            self.solve()
        found = []
        for key, node in self.solution.items():
            match = CELL.match(key)
            if not match or match.group("sheet").upper() in HIDDEN_FROM_BUYER:
                continue
            try:
                value = node.value[0, 0]
            except Exception:
                continue
            if isinstance(value, str) and value.startswith("#"):
                found.append("{0}!{1} {2}".format(match.group("sheet"), match.group("ref"), value))
        return sorted(set(found))


class Case(object):
    def __init__(self, model, name):
        self.m = model
        self.name = name
        self.failures = []
        model.reset()
        self.blank_example()

    def blank_example(self):
        """Return the workbook to the state of a buyer who has typed nothing."""
        for col in ("A", "C", "D", "F", "G", "K", "N"):
            self.m.clear(REGISTER, "{0}{1}".format(col, EXAMPLE))

    def row(self, row, system, status, year=None, life=None, cost=None, place=None):
        self.m.set(REGISTER, "A{0}".format(row), system)
        self.m.set(REGISTER, "F{0}".format(row), status)
        if year is not None:
            self.m.set(REGISTER, "G{0}".format(row), year)
        if life is not None:
            self.m.set(REGISTER, "K{0}".format(row), life)
        if cost is not None:
            self.m.set(REGISTER, "N{0}".format(row), cost)
        if place is not None:
            self.m.set(REGISTER, "C{0}".format(row), place)

    def reg(self, ref):
        return self.m.get(REGISTER, ref)

    def dash(self, key):
        return self.m.get(DASHBOARD, bp.DASH[key][0])

    def now(self):
        return int(self.m.get(SETUP, "B4"))

    def year_cell(self, row, year):
        first = int(self.m.get(FORECAST, "C4"))
        col = 3 + (year - first)
        if not 3 <= col <= 2 + bp.HORIZON:
            raise ValueError("{0} is outside the forecast window".format(year))
        return self.m.get(FORECAST, "{0}{1}".format(get_column_letter(col), row))

    def summary(self, row, year):
        first = int(self.m.get(FORECAST, "C4"))
        return self.m.get(FORECAST, "{0}{1}".format(
            get_column_letter(3 + (year - first)), row))

    def check(self, label, got, want, tolerance=0.51):
        if isinstance(want, (int, float)) and not isinstance(want, bool):
            try:
                ok = abs(float(got) - float(want)) <= tolerance
            except (TypeError, ValueError):
                ok = False
        else:
            ok = got == want
        if not ok:
            self.failures.append("{0}: got {1!r}, wanted {2!r}".format(label, got, want))
        return ok

    def assert_true(self, label, condition):
        if not condition:
            self.failures.append(label)
        return condition

    def is_blank(self, label, value):
        return self.assert_true("{0}: got {1!r}".format(label, value), value in (None, ""))

    def no_errors(self, label="no error cells"):
        found = self.m.errors()
        return self.assert_true("{0}, found {1}".format(label, found[:6]), not found)


# ------------------------------------------------------------------------ cases

def case_1_empty(case):
    case.m.clear(SETUP, "B5")
    case.no_errors()
    case.check("systems tracked", case.dash("systems"), 0)
    case.check("funded %", case.dash("funded"), 0)
    case.check("gross monthly", case.dash("gross"), 0)
    case.check("this year's spend", case.summary(bp.ROW_SPEND, case.now()), 0)
    case.is_blank("largest upcoming item is blank", case.dash("largest"))

    # The trap the spec's own formula falls into: Known, with no year typed yet,
    # returns 0 rather than blank and every downstream guard sails past it.
    case.row(EXAMPLE, "Water heater - tank", "Known")
    case.is_blank("effective install year stays blank", case.reg("H5"))
    case.is_blank("status stays blank", case.reg("V5"))
    case.check("no phantom spend", case.summary(bp.ROW_SPEND, case.now()), 0)
    case.no_errors("still no error cells")

    # And the same for an IDK option picked before the build year is filled in.
    case.m.set(REGISTER, "F{0}".format(EXAMPLE), "IDK - Assume Original")
    case.is_blank("no install year without a build year", case.reg("H5"))
    case.no_errors("and none from the IDK branch")


def case_2_overdue(case):
    case.m.set(SETUP, "B5", 1990)
    case.row(EXAMPLE, "Water heater - tank", "Known", year=1990, life=20)
    now = case.now()

    case.check("status", case.reg("V5"), "Overdue")
    case.check("age", case.reg("P5"), now - 1990)
    case.check("years remaining", case.reg("Q5"), 20 - (now - 1990))
    case.check("next replacement clamps to this year", case.reg("R5"), now)

    cost = case.reg("O5")
    case.check("future cost is today's cost", case.reg("S5"), cost, 0.01)
    case.check("monthly set-aside is the whole thing", case.reg("U5"), cost, 0.01)
    case.check("ideal accrued is the whole thing", case.reg("T5"), cost, 0.01)

    # The bug this product exists to not have. Cycling from the install year
    # instead of from column R skips an overdue item to the far side of its next
    # full lifespan, so the most overdue thing in the house contributes nothing
    # to the next ten years of spend.
    case.check("fires this year", case.year_cell(EXAMPLE_FC, now), cost, 0.01)
    case.check("annual spend picks it up", case.summary(bp.ROW_SPEND, now), cost, 0.01)
    case.check("dashboard shows it due this year", case.dash("this_year"), cost, 0.01)
    case.check("overdue count", case.dash("overdue"), 1)
    case.assert_true("does not fire again next year", not case.year_cell(EXAMPLE_FC, now + 1))
    case.check("repeats one lifespan later", case.year_cell(EXAMPLE_FC, now + 20),
               cost * (1.03 ** 20), 1.0)


def case_3_repeat(case):
    case.m.set(SETUP, "B9", 0)  # zero inflation, so the arithmetic is checkable by hand
    case.row(EXAMPLE, "Water heater - tank", "Known", year=2020, life=10)
    cost = case.reg("O5")
    for year in (2030, 2040, 2050):
        case.check("fires in {0}".format(year), case.year_cell(EXAMPLE_FC, year), cost, 0.01)
    for year in (2026, 2029, 2031, 2039, 2049):
        case.assert_true("silent in {0}".format(year), not case.year_cell(EXAMPLE_FC, year))
    case.check("three replacements in thirty years",
               case.summary(bp.ROW_CUMULATIVE, 2055), cost * 3, 0.01)


def case_4_idk_engine(case):
    case.m.set(SETUP, "B5", 1965)
    case.row(EXAMPLE, "Roof - metal", "IDK - Assume On Schedule", life=25)
    now = case.now()
    expected = 1965 + ((now - 1965) // 25) * 25
    case.check("dates from the last cycle, not the build year", case.reg("H5"), expected)
    case.assert_true("and that is not the build year", expected != 1965)
    case.check("confidence", case.reg("I5"), "Estimated")

    case.m.set(REGISTER, "F{0}".format(EXAMPLE), "IDK - Assume Original")
    case.check("original dates to the build year", case.reg("H5"), 1965)
    case.check("and reads overdue", case.reg("V5"), "Overdue")

    case.m.set(REGISTER, "F{0}".format(EXAMPLE), "IDK - Assume Mid-Life")
    case.check("mid-life is half a lifespan back", case.reg("H5"), now - 13)
    case.check("which is half used up", case.reg("P5"), 13)

    case.m.set(REGISTER, "F{0}".format(EXAMPLE), "Known")
    case.m.set(REGISTER, "G{0}".format(EXAMPLE), 2018)
    case.check("known wins", case.reg("H5"), 2018)
    case.check("confidence", case.reg("I5"), "High")

    # A house younger than its own systems is nonsense, and it used to be a #NUM!
    case.m.set(SETUP, "B5", 2030)
    case.m.set(REGISTER, "F{0}".format(EXAMPLE), "IDK - Assume On Schedule")
    case.check("a build year in the future does not blow up", case.reg("H5"), 2030)
    case.no_errors("and raises no error")


def case_5_overrides(case):
    case.m.set(SETUP, "B9", 0)
    case.row(EXAMPLE, "HVAC condenser - outdoor unit", "Known", year=2020)
    now = case.now()
    default_life = case.reg("J5")
    default_cost = case.reg("M5")
    case.check("lifespan used starts at the default", case.reg("L5"), default_life)
    case.check("cost used starts at the default", case.reg("O5"), default_cost)
    first_year = case.reg("R5")

    case.m.set(REGISTER, "K{0}".format(EXAMPLE), 20)
    case.m.set(REGISTER, "N{0}".format(EXAMPLE), 9000)
    case.check("lifespan used follows the override", case.reg("L5"), 20)
    case.check("cost used follows the override", case.reg("O5"), 9000)
    case.check("age is unchanged", case.reg("P5"), now - 2020)
    case.check("years remaining recalculated", case.reg("Q5"), 20 - (now - 2020))
    case.check("next replacement recalculated", case.reg("R5"), 2040)
    case.check("future cost recalculated", case.reg("S5"), 9000, 0.01)
    case.check("ideal accrued recalculated", case.reg("T5"), 9000 * (now - 2020) / 20.0, 1.0)
    case.check("monthly recalculated", case.reg("U5"),
               9000 / ((20 - (now - 2020)) * 12.0), 0.01)
    case.check("status recalculated", case.reg("V5"), "OK")
    case.check("forecast follows too", case.year_cell(EXAMPLE_FC, 2040), 9000, 0.01)
    case.assert_true("and vacated the old year",
                     not case.year_cell(EXAMPLE_FC, int(first_year)))
    case.check("dashboard follows", case.dash("value"), 9000, 0.01)

    case.m.clear(REGISTER, "K{0}".format(EXAMPLE))
    case.m.clear(REGISTER, "N{0}".format(EXAMPLE))
    case.check("clearing the override restores the default lifespan",
               case.reg("L5"), default_life)
    case.check("and the default cost", case.reg("O5"), default_cost)


def case_6_region(case):
    case.row(EXAMPLE, "Roof - metal", "Known", year=2020)          # exposure driven
    case.row(SECOND, "Dishwasher", "Known", year=2020, place="kitchen")  # not
    case.m.set(SETUP, "B6", "Gulf Coast")
    base_roof_cost = case.reg("M5")
    base_roof_life = case.reg("J5")
    base_dish_life = case.reg("J6")
    base_total = case.summary(bp.ROW_CUMULATIVE, 2055)
    case.check("gulf coast cost factor", case.m.get(SETUP, "B7"), 1.0, 0.001)
    case.check("gulf coast life factor", case.m.get(SETUP, "B8"), 1.0, 0.001)

    case.m.set(SETUP, "B6", "Northeast")
    case.check("cost factor", case.m.get(SETUP, "B7"), 1.25, 0.001)
    case.check("life factor", case.m.get(SETUP, "B8"), 1.15, 0.001)
    case.check("roof cost moved", case.reg("M5"), round(base_roof_cost * 1.25))
    case.check("roof life moved", case.reg("J5"), round(base_roof_life * 1.15))
    case.check("dishwasher life did not", case.reg("J6"), base_dish_life)
    case.assert_true("and the whole forecast moved with it",
                     case.summary(bp.ROW_CUMULATIVE, 2055) != base_total)

    case.m.set(SETUP, "B6", "Gulf Coast")
    case.check("switching back restores the cost", case.reg("M5"), base_roof_cost)
    case.check("and the lifespan", case.reg("J5"), base_roof_life)


def case_7_inflation(case):
    case.row(EXAMPLE, "Refrigerator", "Known", year=2024)
    case.m.set(SETUP, "B9", 0)
    cost = case.reg("O5")
    year = int(case.reg("R5"))
    case.check("future cost equals today's cost", case.reg("S5"), cost, 0.001)
    case.check("and so does the forecast cell", case.year_cell(EXAMPLE_FC, year), cost, 0.001)

    case.m.set(SETUP, "B9", 0.03)
    remaining = case.reg("Q5")
    case.check("3 percent compounds", case.reg("S5"), cost * (1.03 ** remaining), 0.5)
    case.check("and the forecast agrees", case.year_cell(EXAMPLE_FC, year),
               cost * (1.03 ** (year - case.now())), 0.5)
    case.check("register and forecast agree with each other",
               case.reg("S5"), case.year_cell(EXAMPLE_FC, year), 0.01)


def case_8_shortfall(case):
    case.row(EXAMPLE, "Roof - asphalt shingle, architectural", "Known", year=2016)
    case.m.set(SETUP, "B10", 0)
    case.m.set(SETUP, "B11", 5)
    first = int(case.m.get(FORECAST, "C4"))
    years = [first + i for i in range(bp.HORIZON)]

    negatives = [y for y in years if (case.summary(bp.ROW_BALANCE, y) or 0) < 0]
    case.assert_true("the balance goes negative", bool(negatives))
    flags = [case.summary(bp.ROW_FLAG, y) for y in negatives]
    case.assert_true("and says SHORTFALL wherever it does",
                     bool(flags) and all(f == "SHORTFALL" for f in flags))
    clean = [case.summary(bp.ROW_FLAG, y) for y in years if y not in negatives]
    case.assert_true("and says nothing where it does not", all(f in (None, "") for f in clean))

    case.m.set(SETUP, "B11", 4000)
    still = [y for y in years if (case.summary(bp.ROW_BALANCE, y) or 0) < 0]
    case.assert_true("a real contribution clears it, {0}".format(still[:4]), not still)
    case.assert_true("and the gap reads as ahead rather than behind", case.dash("gap") < 0)


def case_9_duplicates(case):
    case.row(EXAMPLE, "HVAC condenser - outdoor unit", "Known", year=2012, place="Upstairs unit")
    case.row(SECOND, "HVAC condenser - outdoor unit", "Known", year=2022, place="Downstairs unit")
    case.check("two rows counted", case.dash("systems"), 2)
    case.check("older one is overdue", case.reg("V5"), "Overdue")
    case.check("newer one is not", case.reg("V6"), "OK")
    case.assert_true("different replacement years", case.reg("R5") != case.reg("R6"))
    case.check("forecast carries the location through",
               case.m.get(FORECAST, "B{0}".format(EXAMPLE_FC)), "Upstairs unit")
    case.check("and the second one",
               case.m.get(FORECAST, "B{0}".format(SECOND_FC)), "Downstairs unit")
    case.check("tier pulled for both", case.reg("W6"), "1 - Critical")
    case.check("both counted in tier 1", case.dash("tier_one"), 2)
    case.check("replacement value is the pair", case.dash("value"),
               case.reg("O5") + case.reg("O6"), 0.01)

    case.m.set(REGISTER, "X{0}".format(SECOND), "No")
    total_with = case.summary(bp.ROW_CUMULATIVE, 2055)
    case.m.set(REGISTER, "X{0}".format(EXAMPLE), "No")
    case.assert_true("excluding one drops it out of the forecast only",
                     case.summary(bp.ROW_CUMULATIVE, 2055) < total_with)


def case_11_quick_check(case):
    """The sixty second tab: two cells on Setup and it answers."""
    first, last = bp.QC_FIRST, bp.QC_LAST
    roof, heater = first, first + bp.COMMON.index("Water heater - tank")

    case.m.clear(SETUP, "B5")
    case.is_blank("no install year without a build year", case.m.get(QUICK, "D{0}".format(roof)))
    case.assert_true("and the headline says what to do",
                     "build year" in str(case.m.get(QUICK, "A{0}".format(bp.QC_SUMMARY))))
    case.no_errors("empty Quick Check is clean")

    case.m.set(SETUP, "B5", 1975)
    now = case.now()
    life = case.m.get(QUICK, "B{0}".format(roof))
    expected = 1975 + int((now - 1975) // life) * life
    case.check("dates from the last cycle", case.m.get(QUICK, "D{0}".format(roof)), expected)
    case.check("due follows", case.m.get(QUICK, "E{0}".format(roof)), expected + life)
    case.check("years left follows", case.m.get(QUICK, "F{0}".format(roof)),
               expected + life - now)

    # On schedule can never be overdue, which is the whole reason the tab says so.
    statuses = [case.m.get(QUICK, "G{0}".format(r)) for r in range(first, last + 1)]
    case.check("all twelve report", len([s for s in statuses if s]), len(bp.COMMON))
    case.assert_true("nothing is overdue on the generous default",
                     "Overdue" not in statuses)

    case.m.set(QUICK, "C{0}".format(roof), "Never replaced")
    case.check("never replaced dates to the build year",
               case.m.get(QUICK, "D{0}".format(roof)), 1975)
    case.check("and reads overdue", case.m.get(QUICK, "G{0}".format(roof)), "Overdue")
    headline = str(case.m.get(QUICK, "A{0}".format(bp.QC_SUMMARY)))
    case.assert_true("the headline counts it: {0!r}".format(headline),
                     "of 12" in headline and "Already overdue: 1" in headline)

    case.m.set(QUICK, "C{0}".format(heater), "Roughly half way")
    # Excel's ROUND goes half away from zero. Python's round() goes to even, so
    # a 9 year life lands a year apart depending on which one you ask.
    half = math.floor(case.m.get(QUICK, "B{0}".format(heater)) / 2.0 + 0.5)
    case.check("half way is half a lifespan back", case.m.get(QUICK, "D{0}".format(heater)),
               now - half)

    case.m.set(QUICK, "C{0}".format(heater), "Not in my house")
    case.is_blank("skipping empties the row", case.m.get(QUICK, "D{0}".format(heater)))
    case.is_blank("and its status", case.m.get(QUICK, "G{0}".format(heater)))
    case.assert_true("and drops out of the count",
                     "of 11" in str(case.m.get(QUICK, "A{0}".format(bp.QC_SUMMARY))))

    # Same region rule as the register: weather ages a roof and not a dishwasher.
    dish = first + bp.COMMON.index("Dishwasher")
    base_roof = case.m.get(QUICK, "B{0}".format(roof))
    base_dish = case.m.get(QUICK, "B{0}".format(dish))
    case.m.set(SETUP, "B6", "Pacific NW")
    case.check("roof life moved", case.m.get(QUICK, "B{0}".format(roof)),
               round(base_roof * 1.20))
    case.check("dishwasher did not", case.m.get(QUICK, "B{0}".format(dish)), base_dish)
    case.no_errors("and none of that raised an error")


CASES = [
    ("1  empty workbook", case_1_empty),
    ("2  overdue item", case_2_overdue),
    ("3  repeat cycle", case_3_repeat),
    ("4  the IDK engine", case_4_idk_engine),
    ("5  override precedence", case_5_overrides),
    ("6  region switch", case_6_region),
    ("7  inflation sanity", case_7_inflation),
    ("8  negative balance", case_8_shortfall),
    ("9  duplicate systems", case_9_duplicates),
    ("11 quick check", case_11_quick_check),
]


# ------------------------------------------------------- case 10, file structure

def structure_checks(path):
    """Case 10: what the formula engine cannot see, read out of the shipped file."""
    problems = []

    def want(label, got, expected):
        if got != expected:
            problems.append("{0}: got {1!r}, wanted {2!r}".format(label, got, expected))

    wb = load_workbook(path)
    want("tab order", wb.sheetnames, [
        "START HERE", "Setup", "Quick Check", "Systems Register", "Dashboard",
        "30-Year Forecast", "Replacement Log", "Reference Data", "Lists"])
    want("Lists is hidden", wb["Lists"].sheet_state, "hidden")
    for name in wb.sheetnames:
        if name != "Lists":
            want("{0} is visible".format(name), wb[name].sheet_state, "visible")

    reg = wb["Systems Register"]
    ranges = {}
    for dv in reg.data_validations.dataValidation:
        for cells in dv.sqref.ranges:
            ranges[cells.coord.split(":")[0][0]] = dv.formula1
    for col, listname in (("A", "=SystemList"), ("F", "=StatusList"),
                          ("W", "=TierList"), ("X", "=YesNoList")):
        want("column {0} dropdown".format(col), ranges.get(col), listname)

    quick = wb["Quick Check"]
    quick_dv = [dv.formula1 for dv in quick.data_validations.dataValidation]
    want("Quick Check dropdown", quick_dv, ["=QuickList"])
    if quick["C{0}".format(bp.QC_FIRST)].protection.locked:
        problems.append("Quick Check's answer column is locked")
    for col in "ABDEFG":
        if not quick["{0}{1}".format(col, bp.QC_FIRST)].protection.locked:
            problems.append("Quick Check column {0} is unlocked".format(col))
    filled = sum(1 for r in range(bp.QC_FIRST, bp.QC_LAST + 1) if quick["A{0}".format(r)].value)
    want("Quick Check rows", filled, len(bp.COMMON))

    names = set(wb.defined_names)
    for listname in ("SystemList", "StatusList", "TierList", "YesNoList", "RegionList",
                     "QuickList"):
        if listname not in names:
            problems.append("named range {0} is missing".format(listname))

    for name in ("Reference Data", "Lists", "Systems Register", "Setup", "Dashboard",
                 "30-Year Forecast", "Replacement Log", "START HERE", "Quick Check"):
        ws = wb[name]
        if not ws.protection.sheet:
            problems.append("{0} is not protected".format(name))
        if ws.protection.password:
            problems.append("{0} is password protected, which it must never be".format(name))

    # The columns the buyer owns have to be reachable through the protection.
    # The empty ones carry their style on the column rather than the cell, which
    # is what keeps a blank cell genuinely blank.
    formula_columns = set(bp.fx_columns())
    for col in sorted(bp.EDITABLE):
        if col in formula_columns or col == "X":
            locked = reg["{0}{1}".format(col, bp.FIRST + 1)].protection.locked
        else:
            dim = reg.column_dimensions[col]
            locked = dim.protection.locked
            if reg["{0}{1}".format(col, bp.FIRST + 1)].value is not None:
                problems.append("column {0} should be empty for the buyer".format(col))
        if locked:
            problems.append("column {0} is locked but should be editable".format(col))
    for col in ("B", "H", "L", "O", "S", "T", "U", "V"):
        cell = reg["{0}{1}".format(col, bp.FIRST + 1)]
        if not cell.protection.locked:
            problems.append("formula column {0} is unlocked".format(col))

    want("one chart on the forecast", len(wb["30-Year Forecast"]._charts), 1)
    want("status has four conditional formats",
         len(list(wb["Systems Register"].conditional_formatting)), 1)

    ref = wb["Reference Data"]
    rows = sum(1 for r in range(bp.REF_FIRST, bp.REF_LAST + 1) if ref["A{0}".format(r)].value)
    want("reference rows", rows, len(bp.SYSTEMS))
    for r in range(bp.REF_FIRST, bp.REF_LAST + 1):
        if not ref["I{0}".format(r)].value:
            problems.append("{0} has no source note".format(ref["A{0}".format(r)].value))

    fonts = set()
    for name in wb.sheetnames:
        for row in wb[name].iter_rows():
            for cell in row:
                if cell.font and cell.font.name:
                    fonts.add(cell.font.name)
    if fonts - {"Arial"}:
        problems.append("non-default fonts in the file: {0}".format(sorted(fonts - {"Arial"})))

    return problems


CHECKLIST = """
Five things no automation here can reach. Do them once, after importing the
.xlsx into Google Sheets.

  1. File, Import, Replace spreadsheet. Not "Open with": that leaves it as an
     uploaded file rather than a Sheet, and half of this list will look wrong.
  2. Systems Register column A shows a dropdown, and so do F, W and X.
  3. Setup B6 shows the region list, and changing it moves the costs on the
     register and the totals on the Dashboard.
  4. The chart on 30-Year Forecast survived, with both the spend columns and
     the reserve balance line. This is the listing screenshot, so it is the one
     that matters most. If Sheets dropped it, rebuild it there once by hand:
     the data it reads is row 107 and row 110.
  5. Reference Data and Lists are still protected, Lists is still hidden, and
     the warning can be dismissed without a password.

Then share as view only, and hand out the /copy link rather than the /edit one.
Never share edit access to the master.
""".strip()


def main():
    if "--checklist" in sys.argv:
        print(CHECKLIST)
        return 0

    print("building the shipping workbook")
    shipped = bp.build()
    problems = structure_checks(shipped)
    if problems:
        print("FAIL  10 file structure")
        for line in problems:
            print("        {0}".format(line))
    else:
        print("ok    10 file structure")

    print("building a {0} row copy to calculate".format(QA_ROWS))
    bp.configure(QA_ROWS)
    path = os.path.join(tempfile.gettempdir(), "qa-planner.xlsx")
    bp.build(path)
    model = Model(path)

    failed = 1 if problems else 0
    for name, run in CASES:
        case = Case(model, name)
        try:
            run(case)
        except Exception as exc:
            case.failures.append("crashed: {0!r}".format(exc))
        if case.failures:
            failed += 1
            print("FAIL  {0}".format(name))
            for line in case.failures:
                print("        {0}".format(line))
        else:
            print("ok    {0}".format(name))

    print()
    if failed:
        print("{0} of {1} cases failed".format(failed, len(CASES) + 1))
    else:
        print("all {0} cases pass. Now run --checklist for the Google Sheets half."
              .format(len(CASES) + 1))
    return 1 if failed else 0


if __name__ == "__main__":
    sys.exit(main())
