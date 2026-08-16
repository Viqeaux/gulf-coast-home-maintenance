"""Builds the Home Systems Register and Replacement Fund Forecaster.

    python build_planner.py

Writes `product/gulf-coast-reserve-planner.xlsx`, eight tabs, no macros. The
buyer gets that file and a Google Sheets copy link made by importing it once.

**The .xlsx is the master and Sheets is the import, not the other way round.**
Building it here rather than by hand in Sheets is what makes a regional edition
a factor change in `planner_data.py` instead of an afternoon of retyping, and it
is the only version of this product that can be diffed, tagged or rebuilt.

**Every formula has to evaluate in both Excel and Google Sheets**, which rules
out more than it looks like it does. No XLOOKUP, FILTER, SORT, UNIQUE, SEQUENCE,
QUERY or ARRAYFORMULA, none of which survive the round trip cleanly, and no IFS,
SWITCH, TEXTJOIN, MAXIFS or MINIFS, which behave inconsistently on export. What
is left is nested IF, AND, OR, INDEX, MATCH, SUM, SUMIF, COUNTIF, IFERROR, MIN,
MAX, MOD, ROUND, FLOOR, COLUMN, YEAR and TODAY. `formula()` refuses anything
outside that list, so the constraint is enforced rather than remembered.

Two pieces of the model are worth reading before changing anything.

**The IDK engine, column H.** Most competing templates need an install date the
buyer does not have, which is where they lose the sale. This one takes "I don't
know" as an answer and turns the house's own build year into an estimate, three
ways: original, mid-life, or on schedule. Column I then labels the row
`Estimated` so nobody mistakes the output for fact.

**Overdue items fire in the current year, not on the original cycle.** The
obvious way to write the forecast grid is `MOD(year - install, lifespan) = 0`,
and it is wrong in the one case that matters most: a 1990 water heater on a
20 year life was due in 2010, and that test does not come true again until
2030. The most overdue thing in the house contributes nothing to the next ten
years of spend. So column R clamps the next replacement to no earlier than the
current year and the grid cycles from R. Getting this wrong produces a workbook
with no errors in it and quietly wrong numbers, which is worse than a visible
break.
"""

import os
import re
import sys

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from planner_data import (
    COMMON, INSTALL_STATUSES, KIT_ITEMS, QUICK_STATUSES, REGIONS, SYSTEMS, TIERS, YES_NO,
    check_against_kit)

OUT = os.path.join("product", "gulf-coast-reserve-planner.xlsx")

FIRST = 5           # first data row on Systems Register
LAST = 104          # last data row, 100 rows of capacity
HORIZON = 30        # forecast columns built, C through AF
FC_FIRST = 6        # first item row on the forecast grid, maps to register row 5
HORIZON_END = "AF"  # last forecast column, 2 + HORIZON
LOG_FIRST, LOG_LAST = 5, 54

# Quick Check: the sixty second answer, and the tab a new owner sees first.
QC_SUMMARY = 7
QC_HEAD = 12
QC_FIRST = QC_HEAD + 1
QC_LAST = QC_FIRST + len(COMMON) - 1

# Reference Data occupies a known block, and every lookup in the workbook is
# bounded to it. Whole-column references are tempting and cost real time in
# Google Sheets, where a hundred INDEX/MATCH pairs over column A is a hundred
# scans of a million rows.
REF_FIRST = 5
REF_LAST = REF_FIRST + len(SYSTEMS) - 1
REGION_FIRST = 5
REGION_LAST = REGION_FIRST + len(REGIONS) - 1


def _derive():
    """Recompute the layout that hangs off the number of register rows."""
    global FC_LAST, ROW_SPEND, ROW_CUMULATIVE, ROW_CONTRIB, ROW_BALANCE, ROW_FLAG
    FC_LAST = FC_FIRST + (LAST - FIRST)
    ROW_SPEND = FC_LAST + 2
    ROW_CUMULATIVE = ROW_SPEND + 1
    ROW_CONTRIB = ROW_SPEND + 2
    ROW_BALANCE = ROW_SPEND + 3
    ROW_FLAG = ROW_SPEND + 4


def configure(rows):
    """Build a smaller register. Only `qa_planner.py` uses this.

    The formulas are identical either way; a hundred rows times thirty years is
    simply more than a pure-Python formula engine will chew through in a sitting.
    Shrinking the row count rather than the model is what keeps the QA run a
    test of the shipped thing.
    """
    global LAST, LOG_LAST
    LAST = FIRST + rows - 1
    LOG_LAST = LOG_FIRST + rows - 1
    _derive()


_derive()

# Where the Dashboard's numbers land. `sheet_dashboard()` checks the labels
# beside them at build time, so moving a line without updating this map fails
# the build rather than quietly pointing the QA harness at the wrong cell.
DASH = {
    "systems": ("B5", "Total systems tracked"),
    "value": ("B6", "Total replacement value (today's $)"),
    "ideal": ("B7", "Ideal accrued reserve"),
    "balance": ("B8", "Your reserve balance"),
    "funded": ("B9", "Funded %"),
    "surplus": ("B10", "Shortfall / surplus"),
    "gross": ("B13", "Gross monthly set-aside"),
    "net": ("B14", "Net recommended monthly"),
    "current": ("B15", "Your current contribution"),
    "gap": ("B16", "Gap"),
    "overdue": ("B19", "Overdue"),
    "due_soon": ("B20", "Due soon"),
    "this_year": ("B21", "Spend due this year"),
    "next_five": ("B22", "Next 5 years"),
    "largest": ("B23", "Largest upcoming item"),
    "tier_one": ("B27", TIERS[0]),
}

# Every function the workbook is allowed to use. The list is the compatibility
# promise: anything not on it either breaks on the .xlsx round trip or behaves
# differently in one of the two apps.
ALLOWED = {
    "IF", "AND", "OR", "INDEX", "MATCH", "SUM", "SUMIF", "SUMIFS", "COUNTIF",
    "COUNTIFS", "IFERROR", "MIN", "MAX", "MOD", "ROUND", "FLOOR", "YEAR",
    "TODAY", "SUMPRODUCT", "COLUMN",
}

INK = "1A1A1A"          # formula results, the default
INPUT_BLUE = "1F4E79"   # safe to edit
PULLED_GREEN = "1E6B45"  # comes from another tab
MUTED = "6B6B6B"
YELLOW = "FFF2CC"       # a key assumption, review it
GRAY = "F2F2F2"         # locked
HEAD = "D9E2EC"
RULE = Side(style="thin", color="B0BEC5")


def formula(text):
    """Return a formula after checking it can survive both apps.

    Catches the two mistakes that are expensive here: an unbalanced paren, which
    makes Excel offer to repair the file and stalls the QA run before it starts,
    and a function outside the compatible set, which produces #NAME? on one
    platform and works fine on the other.
    """
    if text.count("(") != text.count(")"):
        raise SystemExit("unbalanced parens:\n  {0}".format(text))
    for name in re.findall(r"([A-Z][A-Z0-9\.]*)\s*\(", text):
        if name not in ALLOWED:
            raise SystemExit("{0}() is not on the compatible list:\n  {1}".format(name, text))
    return text


def style(cell, *, color=INK, bold=False, fill=None, fmt=None, align=None,
          wrap=False, locked=True, size=None, italic=False):
    cell.font = Font(name="Arial", size=size or 10, bold=bold, italic=italic, color=color)
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        cell.number_format = fmt
    if align or wrap:
        cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.protection = Protection(locked=locked)
    return cell


def put(ws, ref, value, **kw):
    cell = ws[ref]
    cell.value = value
    return style(cell, **kw)


def heading(ws, ref, text, size=14):
    put(ws, ref, text, bold=True, size=size, color="0F2E4A")


def note(ws, ref, text):
    put(ws, ref, text, color=MUTED, italic=True, wrap=True)


def band(ws, row, headers, width=None):
    """Write a header band and return the row it landed on."""
    for i, text in enumerate(headers):
        cell = ws.cell(row=row, column=i + 1, value=text)
        style(cell, bold=True, fill=HEAD, wrap=True, align="left")
        cell.border = Border(bottom=RULE)
    ws.row_dimensions[row].height = width or 30
    return row


def widths(ws, spec):
    for letter, w in spec.items():
        ws.column_dimensions[letter].width = w


def lock(ws, allow_sort=True):
    """Warning-only protection: on, but with no password.

    A buyer who wants to restructure the sheet clicks Unprotect once and is
    never asked for anything. A locked-down workbook generates support email,
    and the protection here is only there to stop an accidental overwrite of a
    formula column.
    """
    ws.protection.sheet = True
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.sort = not allow_sort
    ws.protection.autoFilter = False


# ---------------------------------------------------------------- Reference Data

def sheet_reference(wb):
    ws = wb.create_sheet("Reference Data")
    heading(ws, "A1", "Reference Data")
    note(ws, "A2",
         "Lifespans and costs the register uses when you have not entered your own. "
         "Costs are 2026 dollars, installed, for the Gulf South, before the regional factor "
         "on Setup. Every one of them is a starting point, not a quote: put your own figure "
         "in the register's override column and the whole workbook follows it.")
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 42

    band(ws, 4, ["System", "Category", "Lifespan (yrs)", "Typical range",
                 "Cost today ($)", "Typical range", "Tier", "Exposure driven",
                 "Where the number comes from"])

    row = 5
    for system, category, life, life_range, cost, cost_range, tier, exposure, source in SYSTEMS:
        put(ws, "A{0}".format(row), system)
        put(ws, "B{0}".format(row), category)
        put(ws, "C{0}".format(row), life, align="center")
        put(ws, "D{0}".format(row), life_range, color=MUTED, align="center")
        put(ws, "E{0}".format(row), cost, fmt='"$"#,##0')
        put(ws, "F{0}".format(row), cost_range, color=MUTED, align="center")
        put(ws, "G{0}".format(row), TIERS[tier - 1])
        put(ws, "H{0}".format(row), exposure, align="center")
        put(ws, "I{0}".format(row), source, color=MUTED, wrap=True)
        row += 1
    last = row - 1
    if last != REF_LAST:
        raise SystemExit("reference block ends at {0}, REF_LAST says {1}".format(last, REF_LAST))

    # The region table. One edit here re-prices and re-times the whole workbook,
    # which is the entire reason the regional editions are a config change.
    put(ws, "K3", "Region factors", bold=True, size=11, color="0F2E4A")
    for i, text in enumerate(["Region", "Cost factor", "Lifespan factor", "Why"]):
        cell = ws.cell(row=4, column=11 + i, value=text)
        style(cell, bold=True, fill=HEAD, wrap=True, align="left")
        cell.border = Border(bottom=RULE)
    for i, (name, cost_f, life_f, why) in enumerate(REGIONS):
        r = 5 + i
        put(ws, "K{0}".format(r), name)
        put(ws, "L{0}".format(r), cost_f, fmt="0.00", align="center")
        put(ws, "M{0}".format(r), life_f, fmt="0.00", align="center")
        put(ws, "N{0}".format(r), why, color=MUTED, wrap=True)
    put(ws, "K13", "The lifespan factor only reaches systems marked Exposure driven above. "
                   "Heat, humidity and UV shorten a roof and do nothing at all to a dishwasher.",
        color=MUTED, italic=True, wrap=True)
    ws.merge_cells("K13:N14")

    widths(ws, {"A": 34, "B": 17, "C": 13, "D": 13, "E": 14, "F": 14, "G": 16,
                "H": 13, "I": 52, "J": 3, "K": 15, "L": 12, "M": 14, "N": 44})
    ws.freeze_panes = "A5"
    lock(ws)
    return ws, last


# ------------------------------------------------------------------------ Lists

def sheet_lists(wb):
    ws = wb.create_sheet("Lists")
    headers = ["Systems", "Install status", "Tier", "Yes/No", "Region", "Quick Check"]
    for i, text in enumerate(headers):
        style(ws.cell(row=1, column=i + 1, value=text), bold=True)
    columns = [
        (1, [system for system, *_rest in SYSTEMS]),
        (2, INSTALL_STATUSES),
        (3, TIERS),
        (4, YES_NO),
        (5, [name for name, *_rest in REGIONS]),
        (6, QUICK_STATUSES),
    ]
    for column, values in columns:
        for i, text in enumerate(values):
            style(ws.cell(row=2 + i, column=column, value=text))
    widths(ws, {"A": 34, "B": 26, "C": 18, "D": 8, "E": 16, "F": 22})
    ws.sheet_state = "hidden"
    lock(ws)

    # Named ranges rather than raw cross-sheet references. A direct
    # `Lists!$A$2:$A$50` inside a data validation is legal in Excel and lost by
    # some importers; a name survives both.
    names = {
        "SystemList": "Lists!$A$2:$A${0}".format(1 + len(SYSTEMS)),
        "StatusList": "Lists!$B$2:$B${0}".format(1 + len(INSTALL_STATUSES)),
        "TierList": "Lists!$C$2:$C${0}".format(1 + len(TIERS)),
        "YesNoList": "Lists!$D$2:$D${0}".format(1 + len(YES_NO)),
        "RegionList": "Lists!$E$2:$E${0}".format(1 + len(REGIONS)),
        "QuickList": "Lists!$F$2:$F${0}".format(1 + len(QUICK_STATUSES)),
    }
    for name, ref in names.items():
        wb.defined_names[name] = DefinedName(name, attr_text=ref)
    return ws


# ------------------------------------------------------------------------ Setup

def sheet_setup(wb):
    ws = wb.create_sheet("Setup")
    heading(ws, "A1", "Setup")
    note(ws, "A2", "Blue cells are yours to fill in. Everything else calculates itself.")
    ws.merge_cells("A2:D2")

    rows = [
        (3, "Property nickname", None, None, "input", "@",
         "Only ever appears on this tab. Useful if you keep one of these per property."),
        (4, "Current year", formula("=YEAR(TODAY())"), None, "calc", "0",
         "Reads the date from your computer. Everything else is measured from it."),
        (5, "Home build year", 2005, None, "assume", "0",
         "EXAMPLE VALUE. Change it to your own before you trust anything else in here. "
         "It is the one number that makes the I-don't-know options work, and it is on your "
         "county appraisal record if you do not have it to hand."),
        (6, "Region", "Gulf Coast", "RegionList", "input", "@",
         "Sets both factors below. Gulf Coast is the base the reference numbers were written for."),
        (7, "Regional cost factor", formula(
            "=IFERROR(INDEX('Reference Data'!$L${0}:$L${1},"
            "MATCH($B$6,'Reference Data'!$K${0}:$K${1},0)),1)".format(
                REGION_FIRST, REGION_LAST)), None, "calc", "0.00",
         "Scales every default cost."),
        (8, "Regional lifespan factor", formula(
            "=IFERROR(INDEX('Reference Data'!$M${0}:$M${1},"
            "MATCH($B$6,'Reference Data'!$K${0}:$K${1},0)),1)".format(
                REGION_FIRST, REGION_LAST)), None, "calc", "0.00",
         "Scales default lifespans, but only for the exposure-driven systems marked on "
         "Reference Data. A roof ages on weather. A dishwasher does not."),
        (9, "Annual cost inflation", 0.03, None, "assume", "0.0%",
         "3 percent is a general construction cost assumption. Adjust it if your area is "
         "running hotter. Set it to zero and future cost equals today's cost exactly."),
        (10, "Current reserve balance ($)", 0, None, "input", '"$"#,##0',
         "What is actually in the account today. Zero is a fine answer and the Dashboard "
         "will tell you what that means."),
        (11, "Current monthly contribution ($)", 0, None, "input", '"$"#,##0',
         "What you put in each month right now."),
        (12, "\"Due Soon\" threshold (years)", 2, None, "input", "0",
         "How much warning you want before something is called Due Soon."),
        (13, "Forecast horizon (years)", 30, None, "input", "0",
         "Up to 30. Lower it and the forecast tab shortens to match."),
    ]

    for row, label, value, listname, kind, fmt, why in rows:
        put(ws, "A{0}".format(row), label, bold=True)
        color = {"input": INPUT_BLUE, "calc": INK, "assume": INPUT_BLUE}[kind]
        cell = put(ws, "B{0}".format(row), value, color=color, fmt=fmt,
                   fill=YELLOW if kind == "assume" else None,
                   locked=kind == "calc")
        cell.border = Border(bottom=RULE)
        note(ws, "C{0}".format(row), why)
        ws.merge_cells("C{0}:F{0}".format(row))
        ws.row_dimensions[row].height = 30
        if listname:
            dv = DataValidation(type="list", formula1="={0}".format(listname), allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(cell)

    put(ws, "A15", "Horizon over 30 does nothing: the forecast tab is built 30 columns wide.",
        color=MUTED, italic=True)
    ws.merge_cells("A15:F15")

    widths(ws, {"A": 32, "B": 16, "C": 18, "D": 18, "E": 18, "F": 18})
    lock(ws)
    return ws


# ------------------------------------------------------------------ Quick Check

def sheet_quick(wb):
    """Twelve systems, two inputs, one answer.

    This exists because the register is a thirty minute job and nobody buys a
    spreadsheet to spend thirty minutes before it tells them anything. Fill in
    two cells on Setup and this tab says what is already past due, which is the
    whole product demonstrated before any work is done. Then it points at the
    register, which is where the money math lives.

    It was a separate HTML download until 2026-08-16. That could not be relied
    on to open on a phone, and a tool that will not open is worth nothing, so it
    moved in here where it opens anywhere the workbook does.
    """
    ws = wb.create_sheet("Quick Check")
    heading(ws, "A1", "Quick Check")
    note(ws, "A2",
         "Two cells on the Setup tab and this answers. It covers the twelve things almost "
         "every house has. When you want the other 37, and what any of it costs, that is the "
         "Systems Register.")
    ws.merge_cells("A2:G2")
    ws.row_dimensions[2].height = 32

    put(ws, "A4", "Home build year", bold=True)
    put(ws, "B4", formula("=IF(Setup!$B$5=\"\",\"\",Setup!$B$5)"), color=PULLED_GREEN, fmt="0")
    note(ws, "C4", "Set it on Setup. Nothing here works without it.")
    put(ws, "A5", "Region", bold=True)
    put(ws, "B5", formula("=Setup!$B$6"), color=PULLED_GREEN)
    note(ws, "C5", "Also on Setup. It moves the weather-driven lifespans.")

    counts = "$G${0}:$G${1}".format(QC_FIRST, QC_LAST)
    headline = formula(
        # Phrased to read correctly at any count. "1 of 12 need attention" and
        # "1 are already overdue" are both wrong, and a formula that picks
        # between is and are for two different numbers is not worth writing.
        "=IF(Setup!$B$5=\"\",\"Put your home's build year on the Setup tab and this fills in.\","
        "\"Needs attention now: \"&COUNTIF({0},\"Overdue\")+COUNTIF({0},\"Due Soon\")&"
        "\" of \"&SUMPRODUCT(({0}<>\"\")*1)&\".   Already overdue: \"&"
        "COUNTIF({0},\"Overdue\")&\".\")".format(counts))
    cell = put(ws, "A{0}".format(QC_SUMMARY), headline, bold=True, size=13, color="9C0006")
    cell.alignment = Alignment(vertical="center")
    ws.merge_cells("A{0}:G{0}".format(QC_SUMMARY))
    ws.row_dimensions[QC_SUMMARY].height = 30
    note(ws, "A{0}".format(QC_SUMMARY + 1),
         "Every line below starts at \"Replaced on schedule\", which assumes whoever owned "
         "the house before you kept up with it. That is the most generous of the three "
         "guesses, so this is the best case rather than the likely one. Switch anything you "
         "know has never been touched to \"Never replaced\" and watch it move.")
    ws.merge_cells("A{0}:G{1}".format(QC_SUMMARY + 1, QC_SUMMARY + 2))
    ws.row_dimensions[QC_SUMMARY + 1].height = 28

    band(ws, QC_HEAD, ["System", "Typical life (yrs)", "What you know",
                       "Estimated install year", "Due", "Years left", "Status"], width=32)

    for offset, system in enumerate(COMMON):
        r = QC_FIRST + offset
        lookup = "MATCH($A{0},{1},0)".format(r, ref_range("A"))
        put(ws, "A{0}".format(r), system)
        put(ws, "B{0}".format(r), formula(
            "=IFERROR(ROUND(INDEX({life},{m})*"
            "IF(INDEX({exp},{m})=\"Yes\",Setup!$B$8,1),0),\"\")".format(
                life=ref_range("C"), exp=ref_range("H"), m=lookup)),
            fmt="0", align="center", fill=GRAY)
        put(ws, "C{0}".format(r), QUICK_STATUSES[0], color=INPUT_BLUE, locked=False)
        put(ws, "D{0}".format(r), formula(
            "=IF($C{r}=\"Not in my house\",\"\",IF(Setup!$B$5=\"\",\"\","
            "IF($B{r}=\"\",\"\","
            "IF($C{r}=\"Never replaced\",Setup!$B$5,"
            "IF($C{r}=\"Roughly half way\",MAX(Setup!$B$5,Setup!$B$4-ROUND($B{r}/2,0)),"
            "Setup!$B$5+FLOOR(MAX(Setup!$B$4-Setup!$B$5,0)/$B{r},1)*$B{r})))))".format(r=r)),
            fmt="0", align="center")
        put(ws, "E{0}".format(r), formula(
            "=IF($D{r}=\"\",\"\",$D{r}+$B{r})".format(r=r)), fmt="0", align="center")
        put(ws, "F{0}".format(r), formula(
            "=IF($D{r}=\"\",\"\",$E{r}-Setup!$B$4)".format(r=r)), fmt="0", align="center")
        put(ws, "G{0}".format(r), formula(
            "=IF($D{r}=\"\",\"\",IF($F{r}<=0,\"Overdue\","
            "IF($F{r}<=Setup!$B$12,\"Due Soon\","
            "IF($F{r}<=5,\"Monitor\",\"OK\"))))".format(r=r)), bold=True, align="center")

    dv = DataValidation(type="list", formula1="=QuickList", allow_blank=True)
    dv.error = "Pick one from the list."
    dv.errorTitle = "Not on the list"
    ws.add_data_validation(dv)
    dv.add("C{0}:C{1}".format(QC_FIRST, QC_LAST))

    fills = [("Overdue", "FFC7CE", "9C0006"), ("Due Soon", "FFE0B2", "7F4F00"),
             ("Monitor", "FFF2CC", "7F6000"), ("OK", "C6EFCE", "006100")]
    for text, fill, font in fills:
        ws.conditional_formatting.add(
            "G{0}:G{1}".format(QC_FIRST, QC_LAST),
            CellIsRule(operator="equal", formula=['"{0}"'.format(text)],
                       fill=PatternFill("solid", bgColor=fill),
                       font=Font(name="Arial", size=10, color=font, bold=True)))

    put(ws, "A{0}".format(QC_LAST + 2),
        "That is the fast answer, and it is a guess made from the age of the house. "
        "The Systems Register takes all 49 systems, your own install dates, your own quotes "
        "and your region, and turns them into what to set aside every month. That is the "
        "part worth the half hour.", color=MUTED, italic=True, wrap=True)
    ws.merge_cells("A{0}:G{1}".format(QC_LAST + 2, QC_LAST + 4))

    widths(ws, {"A": 34, "B": 15, "C": 22, "D": 18, "E": 10, "F": 11, "G": 12})
    ws.freeze_panes = "A{0}".format(QC_FIRST)
    lock(ws)
    return ws


# -------------------------------------------------------------- Systems Register

REGISTER_COLUMNS = [
    ("A", "System", 32, None),
    ("B", "Category", 15, None),
    ("C", "Location / Notes", 30, None),
    ("D", "Brand & Model", 20, None),
    ("E", "Serial #", 16, None),
    ("F", "Install date status", 20, None),
    ("G", "Install year (if known)", 12, "0"),
    ("H", "Effective install year", 12, "0"),
    ("I", "Confidence", 11, None),
    ("J", "Default lifespan (yrs)", 12, "0"),
    ("K", "Your lifespan override", 12, "0"),
    ("L", "Lifespan used", 11, "0"),
    ("M", "Default cost today ($)", 14, '"$"#,##0'),
    ("N", "Your cost override ($)", 14, '"$"#,##0'),
    ("O", "Cost used ($)", 13, '"$"#,##0'),
    ("P", "Current age (yrs)", 11, "0"),
    ("Q", "Years remaining", 11, "0"),
    ("R", "Next replacement year", 12, "0"),
    ("S", "Future cost at replacement", 15, '"$"#,##0'),
    ("T", "Ideal accrued reserve", 14, '"$"#,##0'),
    ("U", "Monthly set-aside (gross)", 14, '"$"#,##0'),
    ("V", "Status", 12, None),
    ("W", "Tier", 16, None),
    ("X", "Include in forecast?", 11, None),
]

# Columns the buyer owns. Everything else is a formula and stays locked.
EDITABLE = set("ACDEFGKNWX")
# Columns whose value came from another tab, drawn in green.
PULLED = set("BJMW")


def fx_columns():
    """Which register columns carry a formula."""
    return sorted(register_formulas(FIRST))


def ref_range(col):
    return "'Reference Data'!${0}${1}:${0}${2}".format(col, REF_FIRST, REF_LAST)


def register_formulas(row):
    r = row
    lookup = "MATCH($A{0},{1},0)".format(r, ref_range("A"))
    return {
        "B": formula("=IFERROR(INDEX({1},{0}),\"\")".format(lookup, ref_range("B"))),

        # The IDK engine. Read the module docstring before touching it.
        "H": formula(
            "=IF(OR($A{r}=\"\",$F{r}=\"\",$L{r}=\"\"),\"\","
            "IF($F{r}=\"Known\",IF($G{r}=\"\",\"\",$G{r}),"
            "IF(Setup!$B$5=\"\",\"\","
            "IF($F{r}=\"IDK - Assume Original\",Setup!$B$5,"
            "IF($F{r}=\"IDK - Assume Mid-Life\",MAX(Setup!$B$5,Setup!$B$4-ROUND($L{r}/2,0)),"
            "IF($F{r}=\"IDK - Assume On Schedule\","
            "Setup!$B$5+FLOOR(MAX(Setup!$B$4-Setup!$B$5,0)/$L{r},1)*$L{r},"
            "\"\"))))))".format(r=r)),

        "I": formula("=IF($A{r}=\"\",\"\",IF($F{r}=\"\",\"\","
                     "IF($F{r}=\"Known\",\"High\",\"Estimated\")))".format(r=r)),

        # Default lifespan, with the regional factor applied only where the
        # reference table says weather is what kills it.
        "J": formula(
            "=IFERROR(ROUND(INDEX({life},{m})*"
            "IF(INDEX({exp},{m})=\"Yes\",Setup!$B$8,1),0),\"\")".format(
                life=ref_range("C"), exp=ref_range("H"), m=lookup)),

        "L": formula("=IF($A{r}=\"\",\"\",IF($K{r}<>\"\",$K{r},$J{r}))".format(r=r)),
        "M": formula("=IFERROR(ROUND(INDEX({cost},{m})*Setup!$B$7,0),\"\")".format(
            cost=ref_range("E"), m=lookup)),
        "O": formula("=IF($A{r}=\"\",\"\",IF($N{r}<>\"\",$N{r},$M{r}))".format(r=r)),
        "P": formula("=IF($H{r}=\"\",\"\",Setup!$B$4-$H{r})".format(r=r)),
        "Q": formula("=IF($H{r}=\"\",\"\",$L{r}-$P{r})".format(r=r)),

        # Clamped to no earlier than this year, so an overdue item lands in the
        # forecast now rather than on the far side of its next full cycle.
        "R": formula("=IF($H{r}=\"\",\"\",MAX($H{r}+$L{r},Setup!$B$4))".format(r=r)),

        "S": formula("=IF($H{r}=\"\",\"\",$O{r}*(1+Setup!$B$9)^MAX($Q{r},0))".format(r=r)),
        "T": formula("=IF($H{r}=\"\",\"\",$O{r}*MIN(MAX($P{r}/$L{r},0),1))".format(r=r)),
        "U": formula("=IF($H{r}=\"\",\"\",IF($Q{r}<=0,$S{r},$S{r}/MAX($Q{r}*12,1)))".format(r=r)),
        "V": formula("=IF($H{r}=\"\",\"\",IF($Q{r}<=0,\"Overdue\","
                     "IF($Q{r}<=Setup!$B$12,\"Due Soon\","
                     "IF($Q{r}<=5,\"Monitor\",\"OK\"))))".format(r=r)),

        # A formula and a dropdown in the same cell: it fills itself in from the
        # reference table, and anyone who wants to re-rank a system types over it
        # and gets the list.
        "W": formula("=IFERROR(INDEX({tier},{m}),\"\")".format(tier=ref_range("G"), m=lookup)),
    }


def sheet_register(wb):
    ws = wb.create_sheet("Systems Register")
    heading(ws, "A1", "Systems Register")
    note(ws, "A2",
         "One row per thing in the house that will need replacing. Pick the system, say what "
         "you know about its age, and leave the rest alone. Blue columns are yours: A, C, D, E, "
         "F, G, K, N, W and X.")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    band(ws, 4, [h for _c, h, _w, _f in REGISTER_COLUMNS], width=42)

    # The columns the buyer types into are styled at the column level and left
    # empty, rather than being written cell by cell. A cell that carries a style
    # and no value is materialized in the file, and a materialized empty cell is
    # not reliably the same thing as an absent one once the file has been through
    # another program. Leaving them out keeps blank meaning blank.
    for col in sorted(EDITABLE - set(fx_columns()) - {"X"}):
        fmt = dict((c, f) for c, _h, _w, f in REGISTER_COLUMNS)[col]
        dim = ws.column_dimensions[col]
        dim.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        dim.protection = Protection(locked=False)
        if fmt:
            dim.number_format = fmt
        dim.customFormat = True

    for row in range(FIRST, LAST + 1):
        fx = register_formulas(row)
        for col, _head, _w, fmt in REGISTER_COLUMNS:
            if col not in fx and col != "X":
                continue
            cell = ws["{0}{1}".format(col, row)]
            if col in fx:
                cell.value = fx[col]
            if col == "X":
                # Pre-filled so a row joins the forecast the moment it has an
                # age, and so the column reads as a switch rather than a puzzle.
                cell.value = "Yes"
            color = PULLED_GREEN if col in PULLED else (
                INPUT_BLUE if col in EDITABLE else INK)
            style(cell, color=color, fmt=fmt,
                  fill=GRAY if col in ("J", "M") else None,
                  align="center" if col == "X" else None,
                  locked=col not in EDITABLE)
        ws.row_dimensions[row].height = 15

    validations = [
        ("SystemList", "A"), ("StatusList", "F"), ("TierList", "W"), ("YesNoList", "X"),
    ]
    for listname, col in validations:
        dv = DataValidation(type="list", formula1="={0}".format(listname), allow_blank=True)
        dv.error = "Pick one from the list, or turn off protection to add your own."
        dv.errorTitle = "Not on the list"
        ws.add_data_validation(dv)
        dv.add("{0}{1}:{0}{2}".format(col, FIRST, LAST))

    fills = [("Overdue", "FFC7CE", "9C0006"), ("Due Soon", "FFE0B2", "7F4F00"),
             ("Monitor", "FFF2CC", "7F6000"), ("OK", "C6EFCE", "006100")]
    for text, fill, font in fills:
        ws.conditional_formatting.add(
            "V{0}:V{1}".format(FIRST, LAST),
            CellIsRule(operator="equal", formula=['"{0}"'.format(text)],
                       fill=PatternFill("solid", bgColor=fill), font=Font(name="Arial", size=10, color=font, bold=True)))

    # The example. One row, labeled in the place the buyer is already reading.
    put(ws, "A{0}".format(FIRST), "HVAC condenser - outdoor unit", color=INPUT_BLUE, locked=False)
    put(ws, "C{0}".format(FIRST), "Side yard, north wall   <- EXAMPLE ROW, delete or overwrite",
        color=INPUT_BLUE, locked=False)
    put(ws, "D{0}".format(FIRST), "Carrier 24ABC6", color=INPUT_BLUE, locked=False)
    put(ws, "F{0}".format(FIRST), "IDK - Assume Mid-Life", color=INPUT_BLUE, locked=False)

    widths(ws, {c: w for c, _h, w, _f in REGISTER_COLUMNS})
    ws.freeze_panes = "B5"
    lock(ws)
    return ws


# ------------------------------------------------------------------ 30-Year Forecast

def sheet_forecast(wb):
    ws = wb.create_sheet("30-Year Forecast")
    heading(ws, "A1", "30-Year Forecast")
    note(ws, "A2",
         "What each system costs in the year it actually falls due, in that year's dollars. "
         "Things that repeat inside the window, a nine year water heater for instance, fire "
         "more than once.")
    ws.merge_cells("A2:H2")

    last_col = 2 + HORIZON
    style(ws.cell(row=4, column=1, value="System"), bold=True, fill=HEAD)
    style(ws.cell(row=4, column=2, value="Location"), bold=True, fill=HEAD)
    ws.cell(row=4, column=3).value = formula("=Setup!$B$4")
    for c in range(4, last_col + 1):
        prev = get_column_letter(c - 1)
        ws.cell(row=4, column=c).value = formula(
            "=IF(COLUMN()-2<=Setup!$B$13,{0}4+1,\"\")".format(prev))
    for c in range(3, last_col + 1):
        style(ws.cell(row=4, column=c), bold=True, fill=HEAD, fmt="0", align="center")

    for i, row in enumerate(range(FC_FIRST, FC_LAST + 1)):
        r = FIRST + i
        style(ws.cell(row=row, column=1, value=formula(
            "=IF('Systems Register'!$A{0}=\"\",\"\",'Systems Register'!$A{0})".format(r))),
            color=PULLED_GREEN)
        style(ws.cell(row=row, column=2, value=formula(
            "=IF('Systems Register'!$A{0}=\"\",\"\",'Systems Register'!$C{0})".format(r))),
            color=PULLED_GREEN)
        for c in range(3, last_col + 1):
            letter = get_column_letter(c)
            cell = ws.cell(row=row, column=c, value=formula(
                "=IF(OR({L}$4=\"\",'Systems Register'!$X{r}=\"No\","
                "'Systems Register'!$H{r}=\"\",'Systems Register'!$L{r}=\"\","
                "'Systems Register'!$R{r}=\"\"),0,"
                "IF(AND({L}$4>='Systems Register'!$R{r},"
                "MOD({L}$4-'Systems Register'!$R{r},'Systems Register'!$L{r})=0),"
                "'Systems Register'!$O{r}*(1+Setup!$B$9)^({L}$4-Setup!$B$4),0))".format(
                    L=letter, r=r)))
            style(cell, fmt='"$"#,##0;;""')

    labels = [(ROW_SPEND, "Annual spend"), (ROW_CUMULATIVE, "Cumulative spend"),
              (ROW_CONTRIB, "Annual contribution"), (ROW_BALANCE, "Reserve balance (end of year)"),
              (ROW_FLAG, "Shortfall flag")]
    for row, text in labels:
        style(ws.cell(row=row, column=1, value=text), bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    for c in range(3, last_col + 1):
        letter = get_column_letter(c)
        prev = get_column_letter(c - 1)
        style(ws.cell(row=ROW_SPEND, column=c, value=formula(
            "=SUM({0}{1}:{0}{2})".format(letter, FC_FIRST, FC_LAST))),
            bold=True, fmt='"$"#,##0')
        style(ws.cell(row=ROW_CUMULATIVE, column=c, value=formula(
            "={0}{1}".format(letter, ROW_SPEND) if c == 3
            else "={0}{1}+{2}{3}".format(prev, ROW_CUMULATIVE, letter, ROW_SPEND))),
            fmt='"$"#,##0', color=MUTED)
        style(ws.cell(row=ROW_CONTRIB, column=c, value=formula(
            "=IF({0}$4=\"\",0,Setup!$B$11*12)".format(letter))), fmt='"$"#,##0', color=MUTED)
        style(ws.cell(row=ROW_BALANCE, column=c, value=formula(
            "=Setup!$B$10+{0}{1}-{0}{2}".format(letter, ROW_CONTRIB, ROW_SPEND) if c == 3
            else "={0}{1}+{2}{3}-{2}{4}".format(
                prev, ROW_BALANCE, letter, ROW_CONTRIB, ROW_SPEND))),
            bold=True, fmt='"$"#,##0')
        style(ws.cell(row=ROW_FLAG, column=c, value=formula(
            "=IF({0}{1}<0,\"SHORTFALL\",\"\")".format(letter, ROW_BALANCE))),
            bold=True, color="9C0006", align="center")

    end = get_column_letter(last_col)
    ws.conditional_formatting.add(
        "C{0}:{1}{0}".format(ROW_BALANCE, end),
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", bgColor="FFC7CE"),
                   font=Font(name="Arial", size=10, bold=True, color="9C0006")))
    ws.conditional_formatting.add(
        "C{0}:{1}{2}".format(FC_FIRST, end, FC_LAST),
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", bgColor="FFF2CC")))

    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.title = "Replacement spend against reserve balance"
    chart.y_axis.title = "Dollars"
    chart.height, chart.width = 9, 28
    spend = Reference(ws, min_col=3, max_col=last_col, min_row=ROW_SPEND, max_row=ROW_SPEND)
    years = Reference(ws, min_col=3, max_col=last_col, min_row=4, max_row=4)
    chart.add_data(spend, from_rows=True, titles_from_data=False)
    chart.set_categories(years)
    chart.series[0].tx = None

    line = LineChart()
    balance = Reference(ws, min_col=3, max_col=last_col, min_row=ROW_BALANCE, max_row=ROW_BALANCE)
    line.add_data(balance, from_rows=True, titles_from_data=False)
    line.y_axis.axId = 200
    line.y_axis.title = "Reserve balance"
    line.y_axis.crosses = "max"
    chart += line
    ws.add_chart(chart, "A113")

    widths(ws, {"A": 30, "B": 26})
    for c in range(3, last_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.freeze_panes = "C5"
    lock(ws)
    return ws


# -------------------------------------------------------------------- Dashboard

def sheet_dashboard(wb):
    ws = wb.create_sheet("Dashboard")
    heading(ws, "A1", "Dashboard")
    note(ws, "A2", "Everything here reads from the register. Nothing on this tab is typed in.")
    ws.merge_cells("A2:E2")

    reg = "'Systems Register'"
    fc = "'30-Year Forecast'"
    ideal = "SUM({0}!T{1}:T{2})".format(reg, FIRST, LAST)

    def block(row, title):
        put(ws, "A{0}".format(row), title, bold=True, size=12, color="0F2E4A")
        ws.cell(row=row, column=1).border = Border(bottom=RULE)
        ws.merge_cells("A{0}:C{0}".format(row))
        return row + 1

    def line_item(row, label, value, fmt=None, why=None, big=False):
        put(ws, "A{0}".format(row), label)
        put(ws, "B{0}".format(row), value, fmt=fmt, bold=big, size=12 if big else 10,
            color=PULLED_GREEN)
        if why:
            note(ws, "C{0}".format(row), why)
            ws.merge_cells("C{0}:F{0}".format(row))
        return row + 1

    row = block(4, "Funding status")
    # Counting filled rows is the one place here where the obvious formula is a
    # trap. COUNTIF(range,"<>") and COUNTIF(range,"?*") both read as "count the
    # ones with something in them", and both hand the answer to whatever that
    # program decides an empty criterion or a wildcard means. SUMPRODUCT of a
    # comparison has no criteria string in it at all, so there is nothing left
    # to interpret differently.
    row = line_item(row, "Total systems tracked", formula(
        "=SUMPRODUCT(({0}!A{1}:A{2}<>\"\")*1)".format(reg, FIRST, LAST)), "0")
    row = line_item(row, "Total replacement value (today's $)", formula(
        "=SUM({0}!O{1}:O{2})".format(reg, FIRST, LAST)), '"$"#,##0')
    row = line_item(row, "Ideal accrued reserve", formula("={0}".format(ideal)), '"$"#,##0',
                    "What you would have set aside by now if you had been saving against every "
                    "system since it was installed.")
    row = line_item(row, "Your reserve balance", formula("=Setup!$B$10"), '"$"#,##0')
    funded_row = row
    row = line_item(row, "Funded %", formula(
        "=IFERROR(Setup!$B$10/{0},0)".format(ideal)), "0.0%", big=True)
    row = line_item(row, "Shortfall / surplus", formula(
        "=Setup!$B$10-{0}".format(ideal)), '"$"#,##0', big=True)

    row = block(row + 1, "Contribution guidance")
    gross_row = row
    row = line_item(row, "Gross monthly set-aside", formula(
        "=SUM({0}!U{1}:U{2})".format(reg, FIRST, LAST)), '"$"#,##0',
        "Every system funded from scratch, ignoring the money you already have.")
    net_row = row
    row = line_item(row, "Net recommended monthly", formula(
        "=MAX(0,(SUM({0}!C{1}:L{1})-Setup!$B$10)/120)".format(fc, ROW_SPEND)), '"$"#,##0',
        "The next ten years of spend, less what is already in the account, spread over 120 "
        "months. This is the number to actually put in the account. Anything already overdue "
        "sits in year one, so a backlog pushes it up.", big=True)
    cur_row = row
    row = line_item(row, "Your current contribution", formula("=Setup!$B$11"), '"$"#,##0')
    row = line_item(row, "Gap", formula("=B{0}-B{1}".format(net_row, cur_row)), '"$"#,##0',
                    "Positive means you are behind. Negative means you are ahead.", big=True)

    row = block(row + 1, "What needs attention")
    row = line_item(row, "Overdue", formula(
        "=COUNTIF({0}!V{1}:V{2},\"Overdue\")".format(reg, FIRST, LAST)), "0",
        "Past its expected life. Not broken, but on borrowed time.")
    row = line_item(row, "Due soon", formula(
        "=COUNTIF({0}!V{1}:V{2},\"Due Soon\")".format(reg, FIRST, LAST)), "0")
    row = line_item(row, "Spend due this year", formula(
        "={0}!C{1}".format(fc, ROW_SPEND)), '"$"#,##0')
    row = line_item(row, "Next 5 years", formula(
        "=SUM({0}!C{1}:G{1})".format(fc, ROW_SPEND)), '"$"#,##0')
    row = line_item(row, "Largest upcoming item", formula(
        "=IFERROR(INDEX({0}!A{1}:A{2},MATCH(MAX({0}!S{1}:S{2}),{0}!S{1}:S{2},0)),\"\")".format(
            reg, FIRST, LAST)), None)

    row = block(row + 1, "By tier")
    put(ws, "B{0}".format(row), "Count", bold=True, align="center")
    put(ws, "C{0}".format(row), "Replacement value", bold=True, align="center")
    put(ws, "D{0}".format(row), "Monthly set-aside", bold=True, align="center")
    row += 1
    for tier in TIERS:
        put(ws, "A{0}".format(row), tier)
        put(ws, "B{0}".format(row), formula(
            "=COUNTIF({0}!$W${1}:$W${2},\"{3}\")".format(reg, FIRST, LAST, tier)),
            fmt="0", align="center", color=PULLED_GREEN)
        put(ws, "C{0}".format(row), formula(
            "=SUMIF({0}!$W${1}:$W${2},\"{3}\",{0}!$O${1}:$O${2})".format(reg, FIRST, LAST, tier)),
            fmt='"$"#,##0', color=PULLED_GREEN)
        put(ws, "D{0}".format(row), formula(
            "=SUMIF({0}!$W${1}:$W${2},\"{3}\",{0}!$U${1}:$U${2})".format(reg, FIRST, LAST, tier)),
            fmt='"$"#,##0', color=PULLED_GREEN)
        row += 1

    put(ws, "A{0}".format(row + 1),
        "Gross ignores money you have already saved. Net credits it against the next ten years "
        "and is the one to act on. They differ on purpose.", color=MUTED, italic=True, wrap=True)
    ws.merge_cells("A{0}:F{1}".format(row + 1, row + 2))

    ws.conditional_formatting.add(
        "B{0}".format(funded_row),
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color="1E6B45", showValue=True))

    wrong = []
    for key, (ref, label) in DASH.items():
        got = ws["A{0}".format(ref[1:])].value
        if got != label:
            wrong.append("{0} -> {1}: label is {2!r}, DASH says {3!r}".format(key, ref, got, label))
    if wrong:
        raise SystemExit("the Dashboard moved under DASH:\n  " + "\n  ".join(wrong))

    widths(ws, {"A": 34, "B": 16, "C": 20, "D": 20, "E": 18, "F": 18})
    lock(ws)
    return ws, gross_row


# --------------------------------------------------------------- Replacement Log

def sheet_log(wb):
    ws = wb.create_sheet("Replacement Log")
    heading(ws, "A1", "Replacement Log")
    note(ws, "A2",
         "What has actually been done, and what it cost. When you log a replacement here, go "
         "back to the Systems Register and change that row's install year to the year you had "
         "it done. That is what moves it back to the start of its next cycle.")
    ws.merge_cells("A2:H2")
    ws.row_dimensions[2].height = 30

    band(ws, 4, ["Date", "System", "What was done", "Contractor", "Cost",
                 "Warranty expiry", "Notes", "Receipt link"])
    for col in "ABCDEFGH":
        dim = ws.column_dimensions[col]
        dim.font = Font(name="Arial", size=10, color=INPUT_BLUE)
        dim.protection = Protection(locked=False)
        if col == "E":
            dim.number_format = '"$"#,##0'
        dim.customFormat = True
    put(ws, "D{0}".format(LOG_LAST + 2), "Total spent to date", bold=True)
    put(ws, "E{0}".format(LOG_LAST + 2), formula(
        "=SUM(E{0}:E{1})".format(LOG_FIRST, LOG_LAST)), bold=True, fmt='"$"#,##0')

    dv = DataValidation(type="list", formula1="=SystemList", allow_blank=True)
    ws.add_data_validation(dv)
    dv.add("B{0}:B{1}".format(LOG_FIRST, LOG_LAST))

    widths(ws, {"A": 12, "B": 30, "C": 34, "D": 22, "E": 13, "F": 14, "G": 34, "H": 26})
    ws.freeze_panes = "A5"
    lock(ws)
    return ws


# ------------------------------------------------------------------- START HERE

START = [
    ("h", "Home Systems Register & Replacement Fund Forecaster"),
    ("p", "Every expensive thing in your house is already on a clock. This works out what "
          "breaks when, what it will cost in the year it breaks, and what to set aside every "
          "month so the money is there when it does."),
    ("h2", "Do these five things"),
    ("n", "Fill in Setup. The one that matters is the home build year, because it is what makes "
          "the I-don't-know options work."),
    ("n", "Open Quick Check. It will already have an answer for you, from that one number. "
          "That is the sixty second version, and everything below is the real one."),
    ("n", "On Systems Register, pick a system in column A, then say what you know about its age "
          "in column F. Repeat for everything you own. Two air conditioners get two rows."),
    ("n", "Read the Dashboard. Net recommended monthly is the number to act on."),
    ("n", "Look at 30-Year Forecast. Where the reserve balance line goes red is the year the "
          "money runs out at your current contribution."),
    ("h2", "You do not need to know when anything was installed"),
    ("p", "That is the whole point of column F, and it is what most spreadsheets like this get "
          "wrong. Pick the one that fits and the workbook estimates the rest from the age of "
          "the house."),
    ("d", "Known", "You have the date. Put the year in column G."),
    ("d", "IDK - Assume Original", "Treats it as original to the house. Conservative. On an "
          "older house this will read Overdue, and that is the correct signal: go and find out."),
    ("d", "IDK - Assume Mid-Life", "A neutral middle estimate, half its expected life used up."),
    ("d", "IDK - Assume On Schedule", "Optimistic. Assumes whoever owned it before you replaced "
          "it on time, every time."),
    ("p", "Column I labels every estimated row so you never mistake one for a fact. Replace the "
          "guesses as you find the real dates: serial number plates, the county appraisal record, "
          "the previous owner's disclosure."),
    ("h2", "What the colors mean"),
    ("c", "Blue text", "Yours. Type over it."),
    ("c", "Black text", "A calculation. It will come back if you delete it, but leave it alone."),
    ("c", "Green text", "Pulled in from another tab."),
    ("c", "Yellow fill", "An assumption worth reviewing, on Setup."),
    ("c", "Gray fill", "A default from the reference table. Override it in the blue column beside it."),
    ("h2", "Overrides beat defaults, always"),
    ("p", "Columns K and N on the register exist because your house is not the average one. Put "
          "a real quote or a real lifespan in either and everything downstream follows your "
          "number instead of ours. The defaults are a starting point so the workbook does "
          "something useful on the first day, not an estimate of what your contractor will charge."),
    ("h2", "About the numbers"),
    ("p", "Lifespans are Gulf South figures, which run shorter than national averages because "
          "heat, humidity and hard UV age a house faster. They match the Big Ticket Watch List "
          "in the Gulf Coast Home Maintenance Kit. Pick a different region on Setup and both the "
          "costs and the weather-driven lifespans move with it. Costs are installed prices in "
          "2026 dollars and vary widely by market, which is what column N is for."),
    ("h2", "The sheets are protected, and the password is nothing"),
    ("p", "Formula columns are locked so a stray paste cannot quietly break the math. There is "
          "no password: Review, Unprotect Sheet, and it opens. In Google Sheets you get a warning "
          "you can dismiss. It is a seatbelt, not a lock."),
    ("h2", "One thing to do once a year"),
    ("p", "Open it in January, update anything you had replaced, and check the reserve balance "
          "line. Fifteen minutes. That is the whole maintenance burden of this file."),
    ("p", "General planning guidance, not a substitute for a licensed inspector, contractor, or "
          "your insurance policy terms."),
]


def sheet_start(wb):
    ws = wb.create_sheet("START HERE")
    row, step = 1, 0
    for item in START:
        kind, rest = item[0], item[1:]
        if kind == "h":
            put(ws, "A{0}".format(row), rest[0], bold=True, size=16, color="0F2E4A")
            row += 2
        elif kind == "h2":
            row += 1
            cell = put(ws, "A{0}".format(row), rest[0], bold=True, size=12, color="0F2E4A")
            cell.border = Border(bottom=RULE)
            ws.merge_cells("A{0}:C{0}".format(row))
            row += 1
        elif kind == "p":
            put(ws, "A{0}".format(row), rest[0], wrap=True)
            ws.merge_cells("A{0}:C{1}".format(row, row + 1))
            ws.row_dimensions[row].height = 28
            row += 2
        elif kind == "n":
            step += 1
            put(ws, "A{0}".format(row), "{0}.".format(step), bold=True, align="right")
            put(ws, "B{0}".format(row), rest[0], wrap=True)
            ws.merge_cells("B{0}:C{1}".format(row, row + 1))
            ws.row_dimensions[row].height = 26
            row += 2
        else:  # "d" and "c" are both a labeled definition
            put(ws, "A{0}".format(row), rest[0], bold=True,
                color=INPUT_BLUE if kind == "c" else INK)
            put(ws, "B{0}".format(row), rest[1], wrap=True)
            ws.merge_cells("B{0}:C{1}".format(row, row + 1))
            ws.row_dimensions[row].height = 26
            row += 2
    widths(ws, {"A": 24, "B": 60, "C": 40})
    lock(ws)
    return ws


# ----------------------------------------------------------------------- build

def build(out=OUT):
    shared = check_against_kit()

    wb = Workbook()
    wb.remove(wb.active)
    # Arial everywhere, including the cells nothing ever writes to. Left alone,
    # openpyxl stamps them Calibri, which Google Sheets substitutes on import and
    # which is a font we have no reason to be shipping. Index 0 is the record
    # every unstyled cell points at, so it has to be replaced as well as the
    # Normal style.
    base = Font(name="Arial", size=10, color=INK)
    wb._named_styles["Normal"].font = base
    wb._fonts[0] = base

    sheet_start(wb)
    sheet_setup(wb)
    sheet_quick(wb)
    sheet_register(wb)
    sheet_dashboard(wb)
    sheet_forecast(wb)
    sheet_log(wb)
    sheet_reference(wb)
    sheet_lists(wb)

    # Setup before Quick Check because Quick Check reads two cells from it, and
    # Quick Check before the register because it is the answer somebody gets in
    # a minute rather than the one they get in half an hour.
    order = ["START HERE", "Setup", "Quick Check", "Systems Register", "Dashboard",
             "30-Year Forecast", "Replacement Log", "Reference Data", "Lists"]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0

    folder = os.path.dirname(out)
    if folder and not os.path.isdir(folder):
        os.makedirs(folder)
    wb.save(out)

    print("{0}: {1} systems, {2} locked to the kit's Watch List, {3} rows, {4} forecast years"
          .format(out, len(SYSTEMS), shared, LAST - FIRST + 1, HORIZON))
    return out


if __name__ == "__main__":
    sys.exit(0 if build() else 1)
